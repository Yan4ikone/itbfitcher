# =========================================================
# ITB
# Файл: itb.py
#
# pip install openpyxl
#
# EXE:
# pyinstaller --onefile --noconsole --icon=itb.ico itb.py
# =========================================================

import os
import shutil
import traceback

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Protection, Alignment, Font

# =========================================================
# НАСТРОЙКИ
# =========================================================

TEMPLATE_FILE = "Шабломатор.xlsx"
TEMPLATE_EXTRA_FILE = "Шабломатор_доп.xlsx"
CHECK_FILE = "в проверку.xlsx"
CHECK_EXTRA_FILE = "в проверку_доп.xlsx"

OUTPUT_DIR = "распределенные"
MERGED_DIR = "готовый"

SOURCE_START_ROW = 1
OUTPUT_START_ROW = 2
EMPLOYEE_START_ROW = 1


# =========================================================
# СМЕНЫ
# =========================================================

SHIFT_1 = [
    "Боженко",
    "Зубринова",
    "Глебова",
    "Липа",
    "Громыкина",
    "Копылова",
    "Агарагимова",
    "Астахова",
    "Красюкова",
    "Калужская",
    "Макеев",
    "Осокина",
    "Волкивская",
    "Кручинина",
    "Федорова",
    "Катаев",
    "Лымарчук",
    "Лышенко А.",
    "Свободный 1",
    "Свободный 2",
    "Свободный 3"
]

SHIFT_2 = [
    "Кубарев",
    "Ефремова",
    "Ястребова",
    "Захарова",
    "Меньшиков",
    "Дружбин",
    "Елисова",
    "Цыдыпова",
    "Коваленко",
    "Казакова",
    "Куратов",
    "Митрофанова",
    "Монич",
    "Васильева",
    "Сражаева",
    "Лышенко И.",
    "Лымарчук.",
    "Катаев.",
    "Шанина",

    "Свободный 4",
    "Свободный 5",
    "Свободный 6"
]

EXCLUDED_FREE = [
    "Свободный 1",
    "Свободный 2",
    "Свободный 3",
    "Свободный 4",
    "Свободный 5",
    "Свободный 6"
]

CUSTOM_EMPLOYEES = []

ALL_EMPLOYEES = []

for name in SHIFT_1:
    if name not in ALL_EMPLOYEES:
        ALL_EMPLOYEES.append(name)

for name in SHIFT_2:
    if name not in ALL_EMPLOYEES:
        ALL_EMPLOYEES.append(name)


# =========================================================
# ГЛОБАЛЬНЫЕ
# =========================================================

source_file_path = ""
total_rows_count = 0

employee_widgets = {}


# =========================================================
# UI
# =========================================================

BG = "#1e1f22"
BLOCK = "#2b2d31"
ACCENT = "#4e8cff"

GREEN = "#4cff7b"
RED = "#ff6565"

FONT = ("Segoe UI", 10)
FONT_MAIN = ("Segoe UI", 10)
FONT_LABEL = ("Segoe UI Semibold", 11)
FONT_TITLE = ("Segoe UI Semibold", 16)
FONT_BIG = ("Segoe UI Semibold", 40)
FONT_BUTTON = ("Segoe UI Semibold", 11)


# =========================================================
# ROOT
# =========================================================

root = tk.Tk()

root.state("zoomed")
root.minsize(1600, 950)

root.configure(bg=BG)

try:
    root.iconbitmap("itb.ico")
except:
    pass


# =========================================================
# STYLE
# =========================================================

style = ttk.Style()

style.theme_use("clam")

style.configure(
    "TProgressbar",
    thickness=18,
    troughcolor="#2b2d31",
    background=ACCENT
)


# =========================================================
# LOG
# =========================================================

def log(message):

    status_text.insert(
        tk.END,
        message + "\n"
    )

    status_text.see(tk.END)

    root.update()


# =========================================================
# COUNT ROWS
# =========================================================

def count_rows(filepath):

    workbook = load_workbook(filepath)

    sheet = workbook.active

    count = 0

    for row in range(
         SOURCE_START_ROW,
         sheet.max_row + 1
    ):

        description = sheet[f"B{row}"].value

        if description:
            count += 1

    return count


# =========================================================
# READ ROWS
# =========================================================

def read_source_rows(filepath):

    workbook = load_workbook(filepath)

    sheet = workbook.active

    rows = []

    row_id = 1

    # ====================================
    # ЕСТЬ ЛИ ДОП. КОЛОНКА В ИСХОДНИКЕ
    # ====================================

    has_extra = False

    for check_row in range(
            EMPLOYEE_START_ROW,
            min(sheet.max_row + 1, 20)
    ):

        value = sheet[f"C{check_row}"].value

        if value not in [None, ""]:
            has_extra = True
            break

    for row in range(
            EMPLOYEE_START_ROW,
            sheet.max_row + 1
    ):

        link = sheet[f"A{row}"].value

        description = sheet[f"B{row}"].value

        extra_value = None

        if has_extra:
            extra_value = sheet[f"C{row}"].value

        code_107 = sheet[f"H{row}"].value

        if not description:
            continue

        row_data = {

            "id": row_id,

            "link": link,

            "description": description,

            "107": code_107
        }

        # ====================================
        # ДОП КОЛОНКА
        # ====================================

        if has_extra:

            row_data["extra"] = (
                extra_value
            )

        rows.append(row_data)

        row_id += 1

    return rows

# =========================================================
# UPDATE REMAINING
# =========================================================

def update_remaining():

    global total_rows_count

    used = 0

    for name, widgets in employee_widgets.items():

        if not widgets["enabled"].get():
            continue

        try:
            value = int(
                widgets["entry_var"].get()
            )
        except:
            value = 0

        used += value

    remaining = total_rows_count - used

    if remaining < 0:
        remaining = 0

    remaining_big_label.config(
        text=str(remaining)
    )

    if remaining == 0:
        remaining_big_label.config(
            fg=GREEN
        )
    else:
        remaining_big_label.config(
            fg=RED
        )

    remaining_label.config(
        text=f"Осталось: {remaining}"
    )


# =========================================================
# AUTO DISTRIBUTE
# =========================================================

def auto_distribute(changed_name=None):

    global total_rows_count

    active = []
    fixed_total = 0

    # ====================================
    # СОБИРАЕМ АКТИВНЫХ
    # ====================================

    for name, widgets in employee_widgets.items():

        if not widgets["enabled"].get():
            continue

        active.append(name)

    if not active:
        update_remaining()
        return

    # ====================================
    # СЧИТАЕМ ЗАФИКСИРОВАННЫХ
    # ====================================

    for name in active:

        widgets = employee_widgets[name]

        if name == changed_name:
            continue

        if widgets["fixed"].get():

            try:
                fixed_total += int(
                    widgets["entry_var"].get()
                )
            except:
                pass

    # ====================================
    # ТЕКУЩЕЕ ЗНАЧЕНИЕ ИЗМЕНЁННОГО
    # ====================================

    changed_value = 0

    if changed_name:

        try:
            changed_value = int(
                employee_widgets[
                    changed_name
                ]["entry_var"].get()
            )
        except:
            pass

    remaining = (
        total_rows_count
        -
        fixed_total
        -
        changed_value
    )

    if remaining < 0:
        remaining = 0

    # ====================================
    # КОГО АВТОРАСПРЕДЕЛЯЕМ
    # ====================================

    auto_people = []

    for name in active:

        if name == changed_name:
            continue

        widgets = employee_widgets[name]

        if widgets["fixed"].get():
            continue

        auto_people.append(name)

    # ====================================
    # РАСПРЕДЕЛЕНИЕ
    # ====================================

    if auto_people:

        per_person = (
            remaining //
            len(auto_people)
        )

        extra = (
            remaining %
            len(auto_people)
        )

        for i, person in enumerate(auto_people):

            value = per_person

            if i < extra:
                value += 1

            employee_widgets[
                person
            ]["slider_var"].set(value)

            employee_widgets[
                person
            ]["entry_var"].set(str(value))

    update_remaining()


# =========================================================
# SLIDER
# =========================================================

def slider_changed(name, value):

    value = int(float(value))

    current_total = 0

    for employee, widgets in employee_widgets.items():

        if employee == name:
            continue

        if not widgets["enabled"].get():
            continue

        try:
            current_total += int(
                widgets["entry_var"].get()
            )
        except:
            pass

    max_allowed = (
        total_rows_count - current_total
    )

    if value > max_allowed:
        value = max_allowed

    if value < 0:
        value = 0

    employee_widgets[name][
        "slider_var"
    ].set(value)

    employee_widgets[name][
        "entry_var"
    ].set(str(value))

    auto_distribute(name)




# =========================================================
# ENTRY
# =========================================================

def entry_changed(event, name):

    try:

        value = int(
            employee_widgets[
                name
            ]["entry_var"].get()
        )

    except:

        value = 0

    current_total = 0

    for employee, widgets in employee_widgets.items():

        if employee == name:
            continue

        if widgets["fixed"].get():
            continue

        if not widgets["enabled"].get():
            continue

        try:
            current_total += int(
                widgets["entry_var"].get()
            )
        except:
            pass

    max_allowed = (
        total_rows_count - current_total
    )

    if value > max_allowed:
        value = max_allowed

    if value < 0:
        value = 0

    employee_widgets[name][
        "slider_var"
    ].set(value)

    employee_widgets[name][
        "entry_var"
    ].set(str(value))

    auto_distribute(name)


# =========================================================
# TOGGLE EMPLOYEE
# =========================================================

def toggle_employee(name):

    widgets = employee_widgets[name]

    enabled = widgets["enabled"].get()

    state = (
        "normal"
        if enabled
        else "disabled"
    )

    widgets["slider"].config(
        state=state
    )

    widgets["entry"].config(
        state=state
    )

    if not enabled:

        widgets["slider_var"].set(0)
        widgets["entry_var"].set("0")

    auto_distribute()


# =========================================================
# TOGGLE SHIFT
# =========================================================

def toggle_shift(shift_names, shift_var):

    enabled = shift_var.get()

    for name in shift_names:

        if name in EXCLUDED_FREE:
            continue

        widgets = employee_widgets[name]

        widgets["enabled"].set(enabled)

        toggle_employee(name)

    auto_distribute()


# =========================================================
# SELECT FILE
# =========================================================

def select_source_file():

    global source_file_path
    global total_rows_count

    filepath = filedialog.askopenfilename(

        title="Выберите Excel",

        filetypes=[
            ("Excel", "*.xlsx *.xlsm")
        ]
    )

    if not filepath:
        return

    source_file_path = filepath

    total_rows_count = count_rows(
        filepath
    )

    filename_label.config(
        text=os.path.basename(filepath)
    )

    total_rows_label.config(
        text=f"Всего строк: {total_rows_count}"
    )

    for name, widgets in employee_widgets.items():

        widgets["slider"].config(
            to=total_rows_count
        )

    auto_distribute()

    log(f"Файл: {filepath}")
    log(f"Строк: {total_rows_count}")


# =========================================================
# GENERATE
# =========================================================

def generate_files():



    try:

        if not source_file_path:

            messagebox.showerror(
                "Ошибка",
                "Выберите файл"
            )

            return

        if not os.path.exists(
            TEMPLATE_FILE
        ):

            messagebox.showerror(
                "Ошибка",
                f"Нет {TEMPLATE_FILE}"
            )

            return

        rows = read_source_rows(
            source_file_path
        )

        selected = []

        for name, widgets in employee_widgets.items():

            if not widgets["enabled"].get():
                continue

            try:
                count = int(
                    widgets["entry_var"].get()
                )
            except:
                count = 0

            if count <= 0:
                continue

            selected.append({

                "name": name,
                "count": count
            })

        os.makedirs(
            OUTPUT_DIR,
            exist_ok=True
        )

        from datetime import datetime

        timestamp = datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )

        project_folder = os.path.join(
            OUTPUT_DIR,
            timestamp
        )

        os.makedirs(
            project_folder,
            exist_ok=True
        )

        ready_folder = os.path.join(
            project_folder,
            "готовые"
        )

        source_folder = os.path.join(
            project_folder,
            "исходные"
        )

        os.makedirs(
            ready_folder,
            exist_ok=True
        )

        os.makedirs(
            source_folder,
            exist_ok=True
        )

        os.makedirs(
            project_folder,
            exist_ok=True
        )

        # ====================================
        # ПАПКИ
        # ====================================

        source_copy = os.path.join(
            ready_folder,
            "source.xlsx"
        )

        shutil.copy(
            source_file_path,
            source_copy
        )

        shutil.copy(

            source_file_path,

            os.path.join(
                source_folder,
                os.path.basename(
                    source_file_path
                )
            )
        )

        progress["maximum"] = total_rows_count
        progress["value"] = 0

        current_index = 0

        for employee in selected:

            name = employee["name"]
            count = employee["count"]

            log(f"Создание: {name}")

            progress_label.config(
                text=f"Создание: {name}"
            )

            employee_rows = rows[
                current_index:
                current_index + count
            ]

            current_index += count

            output_file = os.path.join(
                project_folder,
                f"{name}.xlsx"
            )

            # ====================================
            # ЕСТЬ ЛИ ДОП КОЛОНКА
            # ====================================

            has_extra_column = any(
                "extra" in row
                for row in employee_rows
            )

            # ====================================
            # ВЫБОР ШАБЛОНА
            # ====================================

            template_file = TEMPLATE_FILE

            if has_extra_column:
                template_file = TEMPLATE_EXTRA_FILE

            shutil.copy(
                template_file,
                output_file
            )

            workbook = load_workbook(
                output_file
            )

            sheet = workbook.active

            current_row = OUTPUT_START_ROW

            # ====================================
            # ЗАПИСЬ СТРОК
            # ====================================

            for row_data in employee_rows:

                # ====================================
                # КОЛОНКИ
                # ====================================

                if has_extra_column:

                    hyper_col = "E"
                    link_col = "F"

                    col107_col = "I"
                    employee_col = "J"

                else:

                    hyper_col = "D"
                    link_col = "E"
                    col107_col = "H"
                    employee_col = "I"

                # ====================================
                # A = ID
                # ====================================

                sheet[f"A{current_row}"] = (
                    row_data["id"]
                )

                # ====================================
                # B = ОПИСАНИЕ
                # ====================================

                sheet[f"B{current_row}"] = (
                    row_data["description"]
                )

                # ====================================
                # ДОП КОЛОНКА
                # ====================================

                extra_value = row_data.get("extra", "")

                if extra_value not in ["", None]:

                    try:
                        extra_value = int(extra_value)
                    except:
                        pass

                sheet[f"D{current_row}"] = extra_value

                # ====================================
                # ССЫЛКА
                # ====================================

                sheet[f"{link_col}{current_row}"] = (
                    row_data["link"]
                )

                # ====================================
                # ГИПЕРССЫЛКА
                # ====================================

                cell = sheet[f"{hyper_col}{current_row}"]

                if row_data["link"]:

                    cell.value = row_data["link"]

                    cell.hyperlink = row_data["link"]

                else:

                    cell.value = ""

                # ====================================
                # 107
                # ====================================

                sheet[f"{col107_col}{current_row}"] = (
                    row_data.get("107")
                )

                # ====================================
                # ФАМИЛИЯ
                # ====================================

                sheet[f"{employee_col}{current_row}"] = (
                    f'=IF(C{current_row}<>"","{name}","")'
                )

                # ====================================
                # СТИЛЬ ГИПЕРССЫЛКИ
                # ====================================

                cell = sheet[
                    f"{hyper_col}{current_row}"
                ]

                sheet.column_dimensions[hyper_col].width = 18

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

                current_row += 1

                progress["value"] += 1

                progress_percent.config(
                    text=f"{int(progress['value'])} / "
                         f"{total_rows_count}"
                )

                root.update()

            sheet.auto_filter.ref = (
                f"A1:I{sheet.max_row}"
                )


            workbook.save(output_file)

        progress_label.config(
            text="Готово"
        )

        messagebox.showinfo(
            "Готово",
            "Файлы созданы"
        )

    except Exception as e:

        traceback.print_exc()

        messagebox.showerror(
            "Ошибка",
            str(e)
        )


# =========================================================
# MERGE
# =========================================================

def merge_files():

    try:

        if not os.path.exists(
            CHECK_FILE
        ):

            messagebox.showerror(
                "Ошибка",
                f"Нет {CHECK_FILE}"
            )

            return

        folder = filedialog.askdirectory(
            title="Папка с файлами"
        )

        if not folder:
            return

        merged_data = {}
        source_path = os.path.join(
            folder,
            "source.xlsx"
        )

        if not os.path.exists(source_path):
            messagebox.showerror(

                "Ошибка",

                "В папке нет source.xlsx"
            )

            return

        source_rows = read_source_rows(
            source_path
        )

        has_extra_column = False

        if len(source_rows) > 0:

            first_row = source_rows[0]

            if "extra" in first_row:
                has_extra_column = True

        source_map = {}

        for row in source_rows:
            source_map[row["id"]] = row

        files = []

        for f in os.listdir(folder):

            lower = f.lower()

            # ====================================
            # ПОДДЕРЖКА XLSX И XLSM
            # ====================================

            if not (
                    lower.endswith(".xlsx")
                    or
                    lower.endswith(".xlsm")
            ):
                continue

            # ====================================
            # ИСКЛЮЧАЕМ СЛУЖЕБНЫЕ
            # ====================================

            if lower in [

                "source.xlsx",

                "source.xlsm",

                "итог.xlsx",

                "итог.xlsm",

                "в проверку.xlsx",

                "в проверку.xlsm"
            ]:
                continue

            files.append(f)

        progress["maximum"] = len(files)
        progress["value"] = 0

        for file_name in files:

            log(
                f"Чтение файла: {file_name}"
            )

            progress_label.config(
                text=f"Чтение: {file_name}"
            )

            full_path = os.path.join(
                folder,
                file_name
            )

            clean_name = (
                file_name
                .replace(".xlsx", "")
                .replace(".XLSX", "")
                .replace(".xlsm", "")
                .replace(".XLSM", "")
            )


            workbook = load_workbook(
                full_path
            )

            sheet = workbook.active

            for row in range(
                OUTPUT_START_ROW,
                sheet.max_row + 1
            ):

                row_id = sheet[f"A{row}"].value

                if row_id is None:
                    continue

                log(
                    f"Строка {row_id} найдена"
                )

                merged_data[row_id] = {

                    "description":
                        sheet[f"B{row}"].value,

                    "code":
                        sheet[f"C{row}"].value,


                    "link":
                         sheet[f"F{row}"].value
                         if has_extra_column
                         else sheet[f"E{row}"].value,

                    "107":
                         sheet[f"I{row}"].value
                         if has_extra_column
                         else sheet[f"H{row}"].value,

                    "employee":
                         sheet[f"J{row}"].value
                         if has_extra_column
                         else sheet[f"I{row}"].value,

                    # Цвет описания
                    "description_fill":
                        copy(sheet[f"B{row}"].fill),

                    "code_fill":
                        copy(sheet[f"C{row}"].fill),


                    "107_fill":
                         copy(
                            sheet[f'I{row}'].fill
                            if has_extra_column
                            else sheet[f'H{row}'].fill
                        ),
                }

            progress["value"] += 1

            root.update()

        check_template = (
            CHECK_EXTRA_FILE
            if has_extra_column
            else CHECK_FILE
        )

        check_book = load_workbook(
            check_template
        )

        check_sheet = check_book.active

        current_row = OUTPUT_START_ROW

        sorted_ids = sorted(
            source_map.keys()
        )

        for row_id in sorted_ids:
            source_data = source_map[row_id]

            data = merged_data.get(
                row_id,
                {}
            )

            # ====================================
            # ИСХОДНЫЕ ДАННЫЕ
            # ====================================

            original_description = (
                source_data["description"]
            )

            original_link = (
                source_data["link"]
            )

            original_code = (
                source_data.get("extra", "")
            )

            # ====================================
            # ИСПРАВЛЕННЫЕ ДАННЫЕ
            # ====================================

            fixed_description = data.get(
                "description",
                ""
            )

            fixed_code = data.get(
                "code",
                ""
            )

            fixed_link = data.get(
                "link",
                original_link
            )

            fixed_107 = data.get(
                "107",
                ""
            )

            fixed_employee = data.get(
                "employee",
                ""
            )

            # ====================================
            # КОЛОНКИ
            # ====================================

            if has_extra_column:

                code_col = "D"

                hyper_col = "F"

                link_col = "G"

                col107_col = "H"

                employee_col = "I"

                check_sheet[f"E{current_row}"] = (
                    original_code
                )

            else:

                code_col = "D"

                hyper_col = "E"

                link_col = "F"

                col107_col = "G"

                employee_col = "H"


            # ====================================
            # ID
            # ====================================

            check_sheet[f"A{current_row}"] = row_id

            # ====================================
            # ИСХОДНОЕ ОПИСАНИЕ
            # ====================================

            check_sheet[f"B{current_row}"] = (
                original_description
            )

            # ====================================
            # ИСПРАВЛЕННОЕ ОПИСАНИЕ
            # ====================================

            check_sheet[f"C{current_row}"] = (
                fixed_description
            )

            # ====================================
            # КОД
            # ====================================

            check_sheet[f"{code_col}{current_row}"] = (
                fixed_code
            )


            # ====================================
            # ГИПЕРССЫЛКА
            # ====================================

            if fixed_link:

                check_sheet[
                    f"{hyper_col}{current_row}"
                ] = (
                    f'=HYPERLINK("{fixed_link}","{fixed_link}")'
                )

            else:

                check_sheet[
                    f"{hyper_col}{current_row}"
                ] = ""

            # ====================================
            # ССЫЛКА
            # ====================================

            check_sheet[
                f"{link_col}{current_row}"
            ] = fixed_link

            # ====================================
            # 107
            # ====================================

            check_sheet[
                f"{col107_col}{current_row}"
            ] = fixed_107

            # ====================================
            # ФАМИЛИЯ
            # ====================================

            check_sheet[
                f"{employee_col}{current_row}"
            ] = fixed_employee



            # ====================================
            # ЗАЛИВКА ОПИСАНИЯ
            # ====================================

            if "description_fill" in data:
                check_sheet[
                    f"C{current_row}"
                ].fill = copy(
                    data["description_fill"]
                )

            # ====================================
            # ЗАЛИВКА КОДА
            # ====================================

            if "code_fill" in data:
                check_sheet[
                    f"{code_col}{current_row}"
                ].fill = copy(
                    data["code_fill"]
                )

            # ====================================
            # ЗАЛИВКА 107
            # ====================================

            if "107_fill" in data:
                check_sheet[
                    f"{col107_col}{current_row}"
                ].fill = copy(
                    data["107_fill"]
                )

            # ====================================
            # СЛЕДУЮЩАЯ СТРОКА
            # ====================================

            current_row += 1

        os.makedirs(
            MERGED_DIR,
            exist_ok=True
        )

        save_path = os.path.join(
            folder,
            "ИТОГ.xlsx"
        )

        if has_extra_column:
            last_col = "I"
        else:
            last_col = "H"

        check_sheet.auto_filter.ref = (
            f"A1:{last_col}{check_sheet.max_row}"
        )

        # ====================================
        # ОЧИСТКА ЛИШНИХ СТРОК В ИТОГЕ
        # ====================================

        for clear_row in range(
                current_row,
                check_sheet.max_row + 1
        ):

            for col in [

                "A",
                "B",
                "C",
                "D",
                "E",
                "F",
                "G",
                "H",
                "I",
                "J"
            ]:
                check_sheet[
                    f"{col}{clear_row}"
                ] = None


        check_book.save(save_path)

        log(
            f"Итоговый файл сохранён: {save_path}"
        )

        messagebox.showinfo(
            "Готово",
            f"Файл сохранён:\n{save_path}"
        )

    except Exception as e:

        traceback.print_exc()

        messagebox.showerror(
            "Ошибка",
            str(e)
        )


# =========================================================
# GUI
# =========================================================

main_container = tk.Frame(
    root,
    bg=BG
)

main_container.pack(
    fill="both",
    expand=True
)

canvas = tk.Canvas(
    main_container,
    bg=BG,
    highlightthickness=0
)

scrollbar = ttk.Scrollbar(
    main_container,
    orient="vertical",
    command=canvas.yview
)

scrollable_frame = tk.Frame(
    canvas,
    bg=BG
)

scrollable_frame.bind(

    "<Configure>",

    lambda e:
    canvas.configure(
        scrollregion=canvas.bbox("all")
    )
)

canvas.create_window(

    (0, 0),

    window=scrollable_frame,

    anchor="nw"
)

canvas.configure(
    yscrollcommand=scrollbar.set
)

canvas.pack(
    side="left",
    fill="both",
    expand=True
)

scrollbar.pack(
    side="right",
    fill="y"
)


def _on_mousewheel(event):

    canvas.yview_scroll(
        int(-1 * (event.delta / 120)),
        "units"
    )


canvas.bind_all(
    "<MouseWheel>",
    _on_mousewheel
)


# =========================================================
# HEADER
# =========================================================

header = tk.Label(

    scrollable_frame,

    bg=BG,

    fg="white",

    font=("Segoe UI", 28, "bold")
)

header.pack(
    pady=0
)


# =========================================================
# TOP
# =========================================================

top_frame = tk.Frame(
    scrollable_frame,
    bg=BLOCK
)

top_frame.pack(
    fill="x",
    padx=15,
    pady=0
)

select_button = tk.Button(

    top_frame,

    text="Выбрать Excel",

    bg=ACCENT,

    fg="white",

    relief="flat",

    font=("Segoe UI", 11, "bold"),

    padx=20,

    pady=10,

    command=select_source_file
)

select_button.pack(
    side="left",
    padx=15,
    pady=15
)

filename_label = tk.Label(

    top_frame,

    text="Файл не выбран",

    bg=BLOCK,

    fg="white",

    font=FONT
)

filename_label.pack(
    side="left",
    padx=15
)


# =========================================================
# INFO
# =========================================================

info_frame = tk.Frame(
    scrollable_frame,
    bg=BLOCK
)

info_frame.pack(
    fill="x",
    padx=15,
    pady=10
)

total_rows_label = tk.Label(

    info_frame,

    text="Всего строк: 0",

    bg=BLOCK,

    fg="white",

    font=("Segoe UI", 14, "bold")
)

total_rows_label.pack(
    anchor="w",
    padx=15,
    pady=10
)

top_remaining_label = tk.Label(

    info_frame,

    text="Осталось: 0",

    bg=BLOCK,

    fg="white",

    font=("Segoe UI", 12)
)

top_remaining_label.pack(
    anchor="w",
    padx=15
)




# =========================================================
# EMPLOYEES
# =========================================================

main_frame = tk.Frame(
    scrollable_frame,
    bg=BG
)

main_frame.pack(
    fill="both",
    expand=True,
    padx=15,
    pady=10
)

main_frame.grid_columnconfigure(0, weight=1)
main_frame.grid_columnconfigure(1, weight=1)
main_frame.grid_columnconfigure(2, weight=0)


left_frame = tk.Frame(
    main_frame,
    bg=BLOCK
)

left_frame.grid(
    row=0,
    column=0,
    sticky="nsew",
    padx=(0, 8)
)

right_frame = tk.Frame(
    main_frame,
    bg=BLOCK
)

right_frame.grid(
    row=0,
    column=1,
    sticky="nsew",
    padx=(8, 0)
)

stats_frame = tk.Frame(
    main_frame,
    bg=BLOCK,
    width=240
)

stats_frame.grid(
    row=0,
    column=2,
    sticky="ns",
    padx=(12, 0)
)

stats_title = tk.Label(

    stats_frame,

    text="Остаток",

    bg=BLOCK,

    fg="white",

    font=("Segoe UI", 16, "bold")
)

stats_title.pack(
    pady=(25, 10)
)

remaining_big_label = tk.Label(

    stats_frame,

    text="0",

    bg=BLOCK,

    fg="#ff6b6b",

    font=("Segoe UI", 42, "bold")
)

remaining_big_label.pack()

remaining_label = tk.Label(

    stats_frame,

    text="Осталось: 0",

    bg=BLOCK,

    fg="white",

    font=("Segoe UI", 12)
)

remaining_label.pack(
    pady=(0, 20)
)

# ====================================
# ДОБАВЛЕНИЕ ФАМИЛИИ
# ====================================

add_frame = tk.Frame(
    scrollable_frame,
    bg=BG
)

add_frame.pack(
    fill="x",
    padx=15,
    pady=(5, 15)
)

employee_name_var = tk.StringVar()

employee_entry = tk.Entry(

    add_frame,

    textvariable=employee_name_var,

    font=("Segoe UI", 11),

    relief="flat"
)

employee_entry.pack(
    side="left",
    padx=(0, 10),
    ipady=4
)

shift_var = tk.StringVar(value="1")

shift_menu = ttk.Combobox(

    add_frame,

    textvariable=shift_var,

    values=["1", "2"],

    width=5,

    state="readonly"
)

shift_menu.pack(
    side="left",
    padx=5
)

def add_employee():

    name = employee_name_var.get().strip()

    if not name:
        return

    shift = shift_var.get()

    if shift == "1":
        SHIFT_1.append(name)
        create_employee_widget(
            left_frame,
            name
        )
    else:
        SHIFT_2.append(name)
        create_employee_widget(
            right_frame,
            name
        )

    CUSTOM_EMPLOYEES.append(name)

    employee_widgets[name]["slider"].config(
        to=total_rows_count
    )

    employee_name_var.set("")

add_btn = tk.Button(

    add_frame,

    text="Добавить",

    command=add_employee,

    bg=ACCENT,

    fg="white",

    relief="flat",

    font=("Segoe UI Semibold", 10),

    padx=15,

    pady=5,

    cursor="hand2"
)

add_btn.pack(
    side="left",
    padx=10
)

def create_employee_widget(parent, employee):

    row_frame = tk.Frame(
        parent,
        bg=BLOCK
    )

    row_frame.pack(
        fill="x",
        padx=10,
        pady=2
    )

    enabled_var = tk.BooleanVar(master=root)
    fixed_var = tk.BooleanVar(master=root)

    checkbox = tk.Checkbutton(

        row_frame,

        variable=enabled_var,

        command=lambda:
        toggle_employee(employee),

        bg=BLOCK,

        fg="white",

        activebackground=BLOCK,

        activeforeground="white",

        selectcolor=BLOCK,

        highlightthickness=0,

        bd=0,

        relief="flat"
    )

    checkbox.pack(side="left")

    name_label = tk.Label(

        row_frame,

        text=employee,

        bg=BLOCK,

        fg="white",

        width=18,

        anchor="w"
    )

    name_label.pack(side="left")

    slider_var = tk.IntVar()

    slider = tk.Scale(

        row_frame,

        from_=0,

        to=0,

        orient="horizontal",

        variable=slider_var,

        bg=BLOCK,

        fg="white",

        troughcolor="#444",

        activebackground=ACCENT,

        highlightthickness=0,

        length=260,

        state="disabled",

        command=lambda value:
        slider_changed(
            employee,
            value
        )
    )

    slider.pack(
        side="left",
        padx=10
    )

    entry_var = tk.StringVar(
        value="0"
    )

    entry = tk.Entry(

        row_frame,

        width=8,

        state="disabled",

        textvariable=entry_var
    )

    entry.pack(
        side="left",
        padx=5
    )

    fixed_cb = tk.Checkbutton(

        row_frame,

        text="Фикс",

        variable=fixed_var,

        bg=BLOCK,

        fg="white",

        activebackground=BLOCK,

        activeforeground="white",

        selectcolor=BLOCK,

        highlightthickness=0,

        bd=0,

        font=("Segoe UI", 9)
    )

    fixed_cb.pack(
        side="left",
        padx=5
    )

    entry.bind(

        "<KeyRelease>",

        lambda event:
        entry_changed(
            event,
            employee
        )
    )

    employee_widgets[employee] = {

        "enabled": enabled_var,

        "fixed": fixed_var,

        "slider_var": slider_var,

        "entry_var": entry_var,

        "slider": slider,

        "entry": entry
    }


header1 = tk.Frame(
    left_frame,
    bg=BLOCK
)

header1.pack(
    fill="x",
    pady=(10, 5)
)

shift1_var = tk.BooleanVar(master=root)

shift1_cb = tk.Checkbutton(

    header1,

    text="Смена 1",

    variable=shift1_var,

    command=lambda:
    toggle_shift(
        SHIFT_1,
        shift1_var
    ),

    bg=BLOCK,

    fg="white",

    activebackground=BLOCK,

    activeforeground="white",

    selectcolor=BLOCK,

    highlightthickness=0,

    bd=0,

    cursor="hand2",

    font=("Segoe UI Semibold", 15)
)

shift1_cb.pack(
    side="left",
    padx=10,
    pady=5
)

for employee in SHIFT_1:
    create_employee_widget(
        left_frame,
        employee
    )


header2 = tk.Frame(
    right_frame,
    bg=BLOCK
)

header2.pack(
    fill="x",
    pady=(10, 5)
)

shift2_var = tk.BooleanVar(master=root)

shift2_cb = tk.Checkbutton(

    header2,

    text="Смена 2",

    variable=shift2_var,

    command=lambda:
    toggle_shift(
        SHIFT_2,
        shift2_var
    ),

    bg=BLOCK,

    fg="white",

    activebackground=BLOCK,

    activeforeground="white",

    selectcolor=BLOCK,

    highlightthickness=0,

    bd=0,

    cursor="hand2",

    font=("Segoe UI Semibold", 15)
)

shift2_cb.pack(
    side="left",
    padx=10,
    pady=5
)

for employee in SHIFT_2:
    if employee in SHIFT_1:
        continue

    create_employee_widget(
        right_frame,
        employee
    )


remaining_label.pack()


button_frame = tk.Frame(
    stats_frame,
    bg=BLOCK
)

button_frame.pack(
    pady=15
)

generate_button = tk.Button(

    button_frame,

    text="Разделить",

    bg="#2d8cff",

    fg="white",

    relief="flat",

    font=("Segoe UI", 11, "bold"),

    padx=20,

    pady=12,

    cursor="hand2",

    bd=0,

    command=generate_files
)

generate_button.pack(
    fill="x",
    pady=5
)

merge_button = tk.Button(

    button_frame,

    text="Собрать",

    bg="#35b56a",

    fg="white",

    relief="flat",

    font=("Segoe UI", 11, "bold"),

    padx=20,

    pady=12,

    cursor="hand2",

    bd=0,

    command=merge_files
)

merge_button.pack(
    fill="x",
    pady=5
)


progress_frame = tk.Frame(
    scrollable_frame,
    bg=BLOCK
)

progress_frame.pack(
    fill="x",
    padx=15,
    pady=10
)

progress_label = tk.Label(

    progress_frame,

    text="Ожидание...",

    bg=BLOCK,

    fg="white",

    font=("Segoe UI", 12, "bold")
)

progress_label.pack(
    anchor="w",
    padx=15,
    pady=10
)

progress = ttk.Progressbar(
    progress_frame,
    mode="determinate"
)

progress.pack(
    fill="x",
    padx=15,
    pady=10
)

progress_percent = tk.Label(

    progress_frame,

    text="0 / 0",

    bg=BLOCK,

    fg="white"
)

progress_percent.pack(
    anchor="w",
    padx=15,
    pady=(0, 10)
)


log_frame = tk.Frame(
    scrollable_frame,
    bg=BLOCK
)

log_frame.pack(
    fill="both",
    expand=True,
    padx=15,
    pady=15
)

status_text = tk.Text(

    log_frame,

    height=10,

    bg="#1b1b1b",

    fg="#e5e5e5",

    insertbackground="white",

    relief="flat",

    font=("Consolas", 10)
)

status_text.pack(
    fill="both",
    expand=True,
    padx=15,
    pady=15
)

log("ITB запущен")

root.mainloop()