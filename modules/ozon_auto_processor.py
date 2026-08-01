import traceback

import openpyxl
import time
import random

from learning.importer import load_learning_history
from parser.cdp_product_parser import CDPProductParser
from engines.decision_engine import DecisionEngine
from modules.decision_logger import DecisionLogger

from pathlib import Path


class OzonAutoProcessor:

    def __init__(
            self,
            excel_path,
            logger=None,
            stats_callback=None,
            limit=None,
            skip_filled=False
    ):
        self.skip_filled = skip_filled
        self.excel_path = excel_path
        self.logger = logger
        self.limit = limit
        self.stats_callback = (stats_callback)
        self.stop_requested = False
        self.total_rows = 0
        self.processed_rows = 0
        self.found_count = 0
        self.not_found_count = 0
        self.learning_buffer = []
        p = Path(excel_path)
        self.result_path = str(p.with_name(f"{p.stem}_RESULT{p.suffix}"))
        self.decision_logger = DecisionLogger()

    def log(self, text):

        print(text)

        if self.logger:
            self.logger(text)

    def stop(self):

        self.stop_requested = True
        self.log("Получена команда остановки...")

    def get_url_from_row(self, ws, row):

        for col in ("D", "E"):

            cell = ws[f"{col}{row}"]

            if cell.hyperlink:
                return cell.hyperlink.target

            value = cell.value

            if (
                    isinstance(value, str)
                    and value.startswith("http")
            ):
                return value

        return None

    def run(self):

        self.start_time = time.time()
        self.log("Открытие Excel...")
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        rows_to_process = []

        for row in range(
                2,
                ws.max_row + 1
        ):

            url = self.get_url_from_row(ws, row)
            current_code = ws[f"C{row}"].value

            if self.skip_filled:

                if current_code not in (
                        None,
                        "",
                        0
                ):
                    continue

            if url:
                rows_to_process.append(row)

        if self.limit:
            rows_to_process = (rows_to_process[:self.limit])

        self.total_rows = len(rows_to_process)
        self.log(f"Всего строк: {self.total_rows}")
        parser = CDPProductParser()
        parser.connect()
        learning_history = load_learning_history(self.excel_path)
        engine = DecisionEngine(learning_history)

        try:
            for row in rows_to_process:
                if self.stop_requested:
                    self.log("Обработка остановлена пользователем")
                    break

                try:

                    self.processed_rows += 1
                    self.process_row(ws, row, parser, engine)

                    if (
                            self.processed_rows
                            % 20
                            == 0
                    ):
                        wb.save(self.result_path)
                        self.log("Файл сохранён")

                except Exception as e:

                    self.log(traceback.format_exc())

                    continue

            wb.save(self.result_path)

        finally:

            parser.disconnect()
        self.print_summary()

    def process_row(self, ws, row, parser, engine):

        url = self.get_url_from_row(ws, row)

        if not url:
            return

        self.log(f"\nСтрока {row}")
        self.log(f"URL: {url}")
        card = parser.parse_url(url)
        excel_name = str(ws[f"B{row}"].value or "").strip()

        if excel_name:
            card.excel_title = excel_name

        result = engine.decide(card)
        ws[f"K{row}"] = "Да" if result.new_product else ""
        ws[f"L{row}"] = "Да" if result.new_dropdown else ""
        self.decision_logger.save(card, result)
        self.learning_buffer.append({
            "url": card.url,
            "title": card.title,
            "description": card.description,
            "product": result.product,
            "code": result.code,
            "material": result.material
        })

        original = ws[f"B{row}"].value

        if result.dropdown:
            ws[f"B{row}"] = result.dropdown
        elif result.product:
            ws[f"B{row}"] = result.product
        else:
            ws[f"B{row}"] = original

        if result.code:

            try:
                ws[f"C{row}"] = int(result.code)
            except ValueError:
                ws[f"C{row}"] = result.code

        if result.code:

            self.found_count += 1
            self.log(f"Описание: {result.dropdown}")
            self.log(f"Товар: {result.product}")
            self.log(f"Материал: {result.material}")
            self.log(f"Код: {result.code}")
            self.log(f"Источник: {result.source}")
            self.log(f"Уверенность: {result.confidence}%")
            if result.review:self.log("⚠ Требуется проверка")

        else:

            self.not_found_count += 1
            self.log(f"Описание: {result.product}")
            self.log("Код не найден")
        self.print_progress()
        delay = random.uniform(2.0,4.0)

        if self.stats_callback:
            self.stats_callback(
                self.total_rows,
                self.processed_rows,
                self.found_count,
                self.not_found_count
            )
        time.sleep(delay)

    def print_progress(self):

        remaining = (
                self.total_rows
                - self.processed_rows
        )
        self.log(
            (
                f"Обработано: "
                f"{self.processed_rows}/"
                f"{self.total_rows} | "
                f"Осталось: "
                f"{remaining}"
            )
        )

    def print_summary(self):

        self.log("\n")
        self.log("=" * 60)
        self.log("ОБРАБОТКА ЗАВЕРШЕНА")
        self.log(f"Всего: {self.total_rows}")
        self.log(f"Найдено: {self.found_count}")
        self.log(f"Не найдено: " f"{self.not_found_count}")

        if self.total_rows:
            percent = round(
                (
                        self.found_count
                        / self.total_rows
                ) * 100,
                2
            )

            self.log(f"Успешность: " f"{percent}%")

        self.log("=" * 60)
        elapsed = round(time.time() - self.start_time)
        self.log(f"Время работы: " f"{elapsed} сек.")

        if self.processed_rows:
            avg = (
                    elapsed
                    / self.processed_rows
            )

            self.log(f"Среднее на товар: " f"{avg:.2f} сек.")

        if self.stats_callback:
            self.stats_callback(
                self.total_rows,
                self.total_rows,
                self.found_count,
                self.not_found_count
            )