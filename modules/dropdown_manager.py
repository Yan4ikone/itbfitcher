import os
import sys

from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from dictionaries.all_dictionaries import MATERIAL_COLORS


sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from dictionaries.products import PRODUCTS
except ImportError:
    print("⚠️ Warning: products.py не найден, выпадающие списки не будут созданы")
    PRODUCTS = {}

from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter



def generate_tnved_codes():
    if not PRODUCTS:
        return []

    unique_codes = set()

    for name, info in PRODUCTS.items():
        if not isinstance(info, dict):
            continue

        code = ""
        for k, v in info.items():
            if "code" in str(k).lower() and "material" not in str(k).lower():
                code = str(v).strip()
                break

        if code and code.isdigit() and code != "0":
            unique_codes.add(int(code))
        for mk, mv in info.items():
            if (
                    "material" in str(mk).lower()
                    and isinstance(mv, dict)
            ):
                for material, mat_code in mv.items():
                    mat_code_str = str(mat_code).strip()

                    if (
                            mat_code_str.isdigit()
                            and mat_code_str != "0"
                    ):
                        unique_codes.add(
                            int(mat_code_str)
                        )
            if mk == "dropdown" and isinstance(mv, dict):
                for variant in mv.get("variants", []):
                    if not isinstance(variant, dict):
                        continue

                    variant_code = str(variant.get("code", "")).strip()

                    if (
                            variant_code.isdigit()
                            and variant_code != "0"
                    ):
                        unique_codes.add(
                            int(variant_code)
                        )
    return sorted(list(unique_codes))


def apply_specific_dropdowns(
        ws,
        desc_col_idx,
        code_col_idx,
        max_row=None
):
    """
    Выпадающие списки берутся непосредственно из PRODUCTS.
    PRODUCTS[product]["dropdown"]:
    """

    if not PRODUCTS:
        return

    limit = max_row if max_row else ws.max_row

    for row in range(2, limit + 1):

        prod_cell = ws.cell(
            row=row,
            column=desc_col_idx
        )

        if not prod_cell.value:
            continue

        prod_name = str(
            prod_cell.value
        ).strip().lower()

        # --------------------------------------------------
        # Ищем товар в PRODUCTS
        # --------------------------------------------------

        product_info = None

        for product_name, info in PRODUCTS.items():

            if not isinstance(info, dict):
                continue

            if str(product_name).strip().lower() in prod_name:
                product_info = info
                break

        if not product_info:
            continue

        dropdown = product_info.get(
            "dropdown"
        )

        if not isinstance(dropdown, dict):
            continue

        variants = dropdown.get(
            "variants",
            []
        )

        if not variants:
            continue

        # --------------------------------------------------
        # Коды
        # --------------------------------------------------

        codes = []

        for item in variants:

            if not isinstance(item, dict):
                continue

            code = str(
                item.get("code", "")
            ).strip()

            if code and code not in codes:
                codes.append(code)

        if not codes:
            continue

        # --------------------------------------------------
        # DataValidation
        # --------------------------------------------------

        formula = '"' + ",".join(codes) + '"'

        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True
        )

        dv.prompt = dropdown.get(
            "title",
            "Выберите вариант"
        )

        dv.showInputMessage = True

        cell = ws.cell(
            row=row,
            column=code_col_idx
        )

        dv.add(cell)
        ws.add_data_validation(dv)

        # --------------------------------------------------
        # Комментарий
        # --------------------------------------------------

        comment_lines = []

        for item in variants:

            if not isinstance(item, dict):
                continue

            code = str(
                item.get("code", "")
            ).strip()

            name = str(
                item.get("name", "")
            ).strip()

            group = str(
                item.get("group", "")
            ).strip()

            if not code:
                continue

            if name and group:
                comment_lines.append(
                    f"{code} - {name} ({group})"
                )
            elif name:
                comment_lines.append(
                    f"{code} - {name}"
                )
            else:
                comment_lines.append(
                    code
                )

        if comment_lines:

            cell.comment = Comment(
                "\n".join(comment_lines),
                "Classifier"
            )

        # --------------------------------------------------
        # Цвет
        #
        # Оставляем старую логику для случаев,
        # когда dropdown фактически представляет
        # один материал.
        # --------------------------------------------------

        materials = set()

        for item in variants:

            if not isinstance(item, dict):
                continue

            name = str(
                item.get("name", "")
            ).strip().lower()

            if name:
                materials.add(name)

        if len(materials) == 1:

            material = next(
                iter(materials)
            )

            color = MATERIAL_COLORS.get(
                material
            )

            if color:

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=color
                )

        break


def apply_dropdowns(wb, ws):
    tnved_codes = generate_tnved_codes()
    if not tnved_codes:
        return

    hidden_sheet_name = "DropdownData"
    if hidden_sheet_name in wb.sheetnames:
        del wb[hidden_sheet_name]

    hidden_ws = wb.create_sheet(hidden_sheet_name)
    hidden_ws.sheet_state = 'hidden'
    header = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    code_col_idx = None
    for idx, col_name in enumerate(header, start=1):
        if "тнвэд" in col_name or "код" in col_name:
            code_col_idx = idx
            break

    if not code_col_idx:
        return

    hidden_ws.cell(row=1, column=1, value="ТН ВЭД Коды")
    for row_idx, code in enumerate(tnved_codes, start=2):
        hidden_ws.cell(row=row_idx, column=1, value=code)
    safe_name = "List_TNVED_Codes"
    last_row = len(tnved_codes) + 1
    ref = f"'{hidden_sheet_name}'!$A$2:$A${last_row}"

    if safe_name in wb.defined_names:
        del wb.defined_names[safe_name]

    defn = DefinedName(safe_name, attr_text=ref)
    wb.defined_names.add(defn)
    dv = DataValidation(
        type="list",
        formula1=f"={safe_name}",
        allow_blank=True
    )
    dv.error = "Выберите код ТН ВЭД из списка"
    dv.errorTitle = "Неверный код"
    dv.prompt = "Выберите код из списка или введите вручную"
    dv.promptTitle = "Код ТН ВЭД"
    col_letter = get_column_letter(code_col_idx)
    dv.add(f"{col_letter}2:{col_letter}10000")
    ws.add_data_validation(dv)

