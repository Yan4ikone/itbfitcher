import json
import os
import queue
import threading
import time
import traceback
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from engines.decision_engine import DecisionEngine
from learning.importer import load_learning_history
from models.card_builder import build_product_card
from processors.pipeline_worker import process_classifier_task

# =============================================================================
# CONFIG
# =============================================================================
CDP_URL = "http://127.0.0.1:9222"
WORKERS = 2
# Отдельный пул процессов для CPU-bound парсинга HTML. Число процессов
# можно держать независимым от WORKERS (числа браузерных потоков) -
# как правило, по числу ядер CPU, минус 1 под остальную систему.
PARSER_PROCESSES = max(1, (os.cpu_count() or 4) - 1)
CLASSIFIER_PROCESSES = 4
NAVIGATION_TIMEOUT = 30_000
MAX_RETRIES = 2
CHECKPOINT_EVERY = 10
DEFAULT_INPUT_FILE = "input.xlsx"
DEFAULT_OUTPUT_FILE = "result.xlsx"
LOG_FILE = "ozon_excel_processor.log"

# Ресурсы, которые не нужны для парсинга (нужен только HTML/JSON) -
# блокировка режет Browser goto время примерно вдвое на тяжёлых
# страницах, т.к. не тратим сеть/CPU на то, что всё равно не читаем.
BLOCKED_RESOURCE_TYPES = ("image", "media", "font", "stylesheet")

# =============================================================================
# RESULT COLUMNS
# =============================================================================
RESULT_COLUMNS = [
    "Parser Status",
    "HTTP Status",
    "Final URL",
    "Title",
    "Description",
    "Price",
    "Currency",
    "SKU",
    "Brand",
    "Main Image",
    "Images",
    "Images Count",
    "Specs",
    "Specs Count",
    "Material",
    "Product",
    "Display Name",
    "TNVED Code",
    "Confidence",
    "Source",
    "Review",
    "Worker",
    "Attempts",
    "Goto Time",
    "HTML Time",
    "Parser Time",
    "Classify Time",
    "Total Time",
    "HTML Size",
    "Blocked Requests",
    "Allowed Requests",
    "Error",
]
# =============================================================================
# LOGGER
# =============================================================================


class Logger:

    def __init__(self, filename: str):

        self.filename = Path(filename)
        self.lock = threading.Lock()
        self.filename.write_text(
            "",
            encoding="utf-8",
        )

    def log(self, message: str):

        with self.lock:

            print(message, flush=True,)

            with self.filename.open(
                "a",
                encoding="utf-8",
            ) as f:

                f.write(message + "\n")
# =============================================================================
# WORK ITEM
# =============================================================================
class WorkItem:

    def __init__(
        self,
        row_number: int,
        url: str,
        attempt: int = 1,
    ):

        self.row_number = row_number
        self.url = url
        self.attempt = attempt


# =============================================================================
# PARSER WORKER (для ProcessPoolExecutor)
#
# ВАЖНО: должна быть top-level функцией модуля (не методом класса) -
# ProcessPoolExecutor должен уметь её импортировать/pickle'ить в
# дочернем процессе. Импорт parse_ozon_html делаем ВНУТРИ функции,
# а не на верхнем уровне модуля - чтобы не тянуть лишние зависимости
# в момент старта дочернего процесса раньше времени.
# =============================================================================


def _parse_html_in_process(html: str) -> dict:

    from parser.ozon_html_parser import parse_ozon_html

    try:
        result = parse_ozon_html(html)

        if not isinstance(result, dict):
            result = {}

        result["__error__"] = None

        return result

    except Exception as exc:

        return {"__error__": str(exc)}
# =============================================================================
# OZON EXCEL PROCESSOR
# =============================================================================

class OzonExcelProcessor:

    def __init__(self, input_file: str, output_file: str | None = None, workers: int = WORKERS):
        self.input_file = Path(input_file)

        if output_file:

            self.output_file = Path(output_file)

        else:

            self.output_file = (self.input_file.parent
                / (
                    self.input_file.stem
                    + "_result.xlsx"
                )
            )

        self.workers = workers
        self.logger = Logger(LOG_FILE)
        self.work_queue = queue.Queue()
        self.result_queue = queue.Queue()
        self.stop_event = threading.Event()
        self.results_lock = threading.Lock()
        self.results = {}
        self.total_urls = 0
        self.completed = 0
        self.success_count = 0
        self.error_count = 0
        self.captcha_count = 0
        self.start_time = None
        self.url_column = None
        self.headers = {}
        self.wb = None
        self.ws = None
        self.writer_thread = None
        self.writer_exception = None
        self.last_save_time = 0
        self.timings = {
            "goto": [],
            "html": [],
            "parser": [],
            "classify": [],
            "total": [],
            "html_size": [],
        }
        self.total_blocked_requests = 0
        self.total_allowed_requests = 0
        self.parser_pool = ProcessPoolExecutor(max_workers=PARSER_PROCESSES)
        self.decision_engine = DecisionEngine(load_learning_history(str(self.input_file)))
        self.phase1_results = []

    # =========================================================================
    # FIND URL COLUMN
    # =========================================================================

    def find_url_column(self):

        possible_names = {
            "url",
            "URL",
            "Url",
            "ссылка",
            "Ссылка",
            "ссылка на товар",
            "Ссылка на товар",
            "ссылка товара",
            "Ссылка товара",
            "product url",
            "product_url",
        }

        for cell in self.ws[1]:

            if cell.value is None:
                continue

            value = str(
                cell.value
            ).strip()

            if value in possible_names:

                self.url_column = (
                    cell.column
                )

                return

        # ---------------------------------------------------------------------
        # Второй вариант — ищем заголовок, содержащий URL/Ссылка
        # ---------------------------------------------------------------------

        for cell in self.ws[1]:

            if cell.value is None:
                continue

            value = str(
                cell.value
            ).strip().lower()

            if (
                "url" in value
                or "ссыл" in value
            ):

                self.url_column = (
                    cell.column
                )

                return

        raise RuntimeError(
            "Не удалось найти колонку URL. "
            "Назови колонку 'URL' или 'Ссылка'."
        )

    # =========================================================================
    # OPEN EXCEL
    # =========================================================================

    def open_excel(self):

        self.logger.log(
            ""
        )

        self.logger.log(
            "=" * 100
        )

        self.logger.log(
            "ОТКРЫТИЕ EXCEL"
        )

        self.logger.log(
            "=" * 100
        )

        self.logger.log(
            f"[EXCEL] Input:  "
            f"{self.input_file}"
        )

        self.logger.log(
            f"[EXCEL] Output: "
            f"{self.output_file}"
        )

        # ---------------------------------------------------------------------
        # Если output уже существует — используем его как checkpoint
        # ---------------------------------------------------------------------

        if self.output_file.exists():

            self.logger.log(
                "[EXCEL] Найден существующий result-файл."
            )

            self.logger.log(
                "[EXCEL] Используем его как checkpoint."
            )

            self.wb = load_workbook(
                self.output_file
            )

        else:

            if not self.input_file.exists():

                raise FileNotFoundError(
                    f"Не найден input: "
                    f"{self.input_file}"
                )

            self.wb = load_workbook(
                self.input_file
            )

        self.ws = self.wb.active

        self.find_url_column()

        self.logger.log(
            f"[EXCEL] URL column: "
            f"{self.url_column}"
        )

        self.logger.log(
            f"[EXCEL] Sheet: "
            f"{self.ws.title}"
        )

        # ---------------------------------------------------------------------
        # Индексы существующих колонок
        # ---------------------------------------------------------------------

        self.headers = {}

        for cell in self.ws[1]:

            if cell.value is not None:

                self.headers[
                    str(cell.value).strip()
                ] = cell.column

        # ---------------------------------------------------------------------
        # Добавляем наши колонки
        # ---------------------------------------------------------------------

        next_column = (
            self.ws.max_column + 1
        )

        for column_name in RESULT_COLUMNS:

            if column_name in self.headers:

                continue

            self.ws.cell(
                row=1,
                column=next_column,
                value=column_name,
            )

            self.headers[
                column_name
            ] = next_column

            next_column += 1

        # ---------------------------------------------------------------------
        # Сохраняем сразу
        # ---------------------------------------------------------------------

        self.wb.save(
            self.output_file
        )

        self.logger.log(
            "[EXCEL] ✓ Workbook готов"
        )

    # =========================================================================
    # READ URLS
    # =========================================================================

    def read_urls(self):

        self.logger.log(
            ""
        )

        self.logger.log(
            "=" * 100
        )

        self.logger.log(
            "ЧТЕНИЕ URL"
        )

        self.logger.log(
            "=" * 100
        )

        count = 0

        skipped = 0

        for row_number in range(
            2,
            self.ws.max_row + 1,
        ):

            cell = self.ws.cell(
                row=row_number,
                column=self.url_column,
            )

            if cell.value is None:

                skipped += 1

                continue

            url = str(
                cell.value
            ).strip()

            if not url:

                skipped += 1

                continue

            # -----------------------------------------------------------------
            # Проверяем существующий статус
            # -----------------------------------------------------------------

            status_column = self.headers.get(
                "Parser Status"
            )

            existing_status = ""

            if status_column:

                value = self.ws.cell(
                    row=row_number,
                    column=status_column,
                ).value

                if value is not None:

                    existing_status = str(
                        value
                    ).strip()

            # -----------------------------------------------------------------
            # Если уже успешно обработано — пропускаем
            # -----------------------------------------------------------------

            if existing_status == "OK":

                skipped += 1

                continue

            # -----------------------------------------------------------------
            # Определяем сайт
            # -----------------------------------------------------------------

            lower_url = url.lower()

            if "ozon.ru" in lower_url:

                self.work_queue.put(
                    WorkItem(
                        row_number=row_number,
                        url=url,
                    )
                )

                count += 1

            elif "wildberries.ru" in lower_url:

                # Пока WB не подключаем.
                self.write_immediate_result(
                    row_number,
                    {
                        "Parser Status":
                            "WB_NOT_IMPLEMENTED",
                        "Error":
                            "WB processor пока не подключён",
                    },
                )

                skipped += 1

            else:

                self.write_immediate_result(
                    row_number,
                    {
                        "Parser Status":
                            "UNSUPPORTED_SITE",
                        "Error":
                            "Неизвестный сайт",
                    },
                )

                skipped += 1

        self.total_urls = count

        self.logger.log(
            f"[EXCEL] Найдено Ozon URL: "
            f"{count}"
        )

        self.logger.log(
            f"[EXCEL] Пропущено: "
            f"{skipped}"
        )

    # =========================================================================
    # IMMEDIATE RESULT
    # =========================================================================

    def write_immediate_result(
        self,
        row_number: int,
        result: dict,
    ):

        for key, value in result.items():

            column = self.headers.get(
                key
            )

            if column:

                self.ws.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )

    # =========================================================================
    # WORKER
    # =========================================================================

    def worker(
        self,
        worker_id: int,
    ):

        self.logger.log(
            ""
        )

        self.logger.log(
            "=" * 100
        )

        self.logger.log(
            f"[WORKER {worker_id}] START"
        )

        self.logger.log(
            "=" * 100
        )

        try:

            # =================================================================
            # ВАЖНО:
            #
            # sync_playwright создаётся ВНУТРИ thread.
            # =================================================================

            with sync_playwright() as p:

                self.logger.log(
                    f"[WORKER {worker_id}] "
                    f"Подключение к CDP..."
                )

                browser = (
                    p.chromium.connect_over_cdp(
                        CDP_URL
                    )
                )

                self.logger.log(
                    f"[WORKER {worker_id}] "
                    f"✓ CDP connected"
                )

                contexts = browser.contexts

                self.logger.log(
                    f"[WORKER {worker_id}] "
                    f"Contexts: "
                    f"{len(contexts)}"
                )

                if not contexts:

                    raise RuntimeError(
                        "CDP не содержит BrowserContext"
                    )

                context = contexts[0]

                # =============================================================
                # СОЗДАЁМ СОБСТВЕННУЮ PAGE
                # =============================================================

                page = context.new_page()

                page.set_default_timeout(
                    NAVIGATION_TIMEOUT
                )

                page.set_default_navigation_timeout(
                    NAVIGATION_TIMEOUT
                )

                # =============================================================
                # БЛОКИРОВКА ТЯЖЁЛЫХ РЕСУРСОВ
                #
                # Парсеру нужен только HTML/JSON. Картинки/шрифты/css/
                # медиа не читаются вообще - их загрузка тратит время
                # goto впустую. Также считаем, сколько запросов
                # заблокировали/пропустили - для профиля.
                # =============================================================

                counters = {"blocked": 0, "allowed": 0}

                # ВАЖНО: ровно ОДИН параметр (route). Playwright сам
                # определяет через интроспекцию сигнатуры, сколько
                # аргументов передать хендлеру - если добавить второй
                # параметр (даже со значением по умолчанию, как было
                # раньше), Playwright решает, что нужно передать
                # (route, request), и в него попадает реальный Request
                # вместо counters. Ошибка молча гасится в except, и
                # блокировка перестаёт работать вообще - именно это
                # произошло в прошлом прогоне (0 заблокировано из 468).
                # counters захватываем через обычное замыкание.
                def route_handler(route):

                    try:
                        if route.request.resource_type in BLOCKED_RESOURCE_TYPES:
                            counters["blocked"] += 1
                            route.abort()
                        else:
                            counters["allowed"] += 1
                            route.continue_()
                    except Exception:
                        try:
                            route.continue_()
                        except Exception:
                            pass

                page.route("**/*", route_handler)

                self.logger.log(
                    f"[WORKER {worker_id}] "
                    f"✓ Page создана (блокировка "
                    f"{', '.join(BLOCKED_RESOURCE_TYPES)} включена)"
                )

                # =============================================================
                # WORK LOOP
                # =============================================================

                while not self.stop_event.is_set():

                    try:

                        item = (
                            self.work_queue.get_nowait()
                        )

                    except queue.Empty:

                        break

                    try:

                        counters["blocked"] = 0
                        counters["allowed"] = 0

                        result = self.process_url(
                            worker_id,
                            page,
                            item,
                        )

                        result["Blocked Requests"] = counters["blocked"]
                        result["Allowed Requests"] = counters["allowed"]

                        with self.results_lock:
                            self.total_blocked_requests += counters["blocked"]
                            self.total_allowed_requests += counters["allowed"]

                        self.result_queue.put(
                            result
                        )

                    except Exception as exc:

                        self.logger.log(
                            f"[WORKER {worker_id}] "
                            f"FATAL:"
                        )

                        self.logger.log(
                            traceback.format_exc()
                        )

                        self.result_queue.put(
                            {
                                "row_number":
                                    item.row_number,
                                "url":
                                    item.url,
                                "Parser Status":
                                    "ERROR",
                                "Worker":
                                    worker_id,
                                "Attempts":
                                    item.attempt,
                                "Error":
                                    str(exc),
                            }
                        )

                    finally:

                        self.work_queue.task_done()

                # =============================================================
                # CLOSE
                # =============================================================

                try:

                    page.close()

                except Exception:

                    pass

                try:

                    browser.close()

                except Exception:

                    pass

        except Exception as exc:

            self.logger.log(
                f"[WORKER {worker_id}] "
                f"✗ WORKER FAILED"
            )

            self.logger.log(
                traceback.format_exc()
            )

            self.writer_exception = exc

        self.logger.log(
            f"[WORKER {worker_id}] STOP"
        )

    # =========================================================================
    # PROCESS URL
    # =========================================================================

    def process_url(
        self,
        worker_id: int,
        page,
        item: WorkItem,
    ):

        row_number = item.row_number

        url = item.url

        attempt = item.attempt

        started = time.perf_counter()

        self.logger.log(
            ""
        )

        self.logger.log(
            "-" * 100
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"ROW {row_number}"
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"URL: {url}"
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"Attempt: {attempt}/{MAX_RETRIES + 1}"
        )

        # =====================================================================
        # GOTO
        # =====================================================================

        goto_start = time.perf_counter()

        response = None

        try:

            response = page.goto(
                url,
                wait_until="domcontentloaded",
                timeout=NAVIGATION_TIMEOUT,
            )

        except Exception as exc:

            goto_time = (
                time.perf_counter()
                - goto_start
            )

            self.logger.log(
                f"[WORKER {worker_id}] "
                f"ROW {row_number} "
                f"✗ GOTO ERROR:"
            )

            self.logger.log(
                repr(exc)
            )

            # -----------------------------------------------------------------
            # RETRY
            # -----------------------------------------------------------------

            if attempt <= MAX_RETRIES:

                self.logger.log(
                    f"[WORKER {worker_id}] "
                    f"ROW {row_number} "
                    f"→ RETRY"
                )

                self.work_queue.put(
                    WorkItem(
                        row_number=row_number,
                        url=url,
                        attempt=attempt + 1,
                    )
                )

                return {
                    "row_number": row_number,
                    "url": url,
                    "Parser Status": "RETRY",
                    "Worker": worker_id,
                    "Attempts": attempt,
                    "Goto Time": goto_time,
                    "Total Time": (
                        time.perf_counter()
                        - started
                    ),
                    "Error": str(exc),
                }

            return {
                "row_number": row_number,
                "url": url,
                "Parser Status": "ERROR",
                "Worker": worker_id,
                "Attempts": attempt,
                "Goto Time": goto_time,
                "Total Time": (
                    time.perf_counter()
                    - started
                ),
                "Error": str(exc),
            }

        goto_time = (
            time.perf_counter()
            - goto_start
        )

        status = (
            response.status
            if response
            else None
        )

        final_url = page.url

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"Status: {status}"
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"Goto: {goto_time:.3f}s"
        )

        # =====================================================================
        # 403 / CAPTCHA
        # =====================================================================

        if status == 403:

            total_time = (
                time.perf_counter()
                - started
            )

            self.logger.log(
                f"[WORKER {worker_id}] "
                f"!!! 403 / CAPTCHA !!!"
            )

            return {
                "row_number": row_number,
                "url": url,
                "Parser Status": "CAPTCHA",
                "HTTP Status": 403,
                "Final URL": final_url,
                "Worker": worker_id,
                "Attempts": attempt,
                "Goto Time": goto_time,
                "Total Time": total_time,
                "Error":
                    "Ozon returned HTTP 403",
            }

        # =====================================================================
        # RETRY 5xx
        # =====================================================================

        if status and status >= 500:

            if attempt <= MAX_RETRIES:

                self.logger.log(
                    f"[WORKER {worker_id}] "
                    f"5xx → RETRY"
                )

                self.work_queue.put(
                    WorkItem(
                        row_number=row_number,
                        url=url,
                        attempt=attempt + 1,
                    )
                )

                return {
                    "row_number": row_number,
                    "url": url,
                    "Parser Status": "RETRY",
                    "HTTP Status": status,
                    "Final URL": final_url,
                    "Worker": worker_id,
                    "Attempts": attempt,
                    "Goto Time": goto_time,
                    "Error":
                        f"HTTP {status}",
                }

        # =====================================================================
        # CONTENT
        # =====================================================================

        html_start = time.perf_counter()

        try:

            html = page.content()

        except Exception as exc:

            html_time = (
                time.perf_counter()
                - html_start
            )

            self.logger.log(
                f"[WORKER {worker_id}] "
                f"CONTENT ERROR: "
                f"{repr(exc)}"
            )

            if attempt <= MAX_RETRIES:

                self.work_queue.put(
                    WorkItem(
                        row_number=row_number,
                        url=url,
                        attempt=attempt + 1,
                    )
                )

                return {
                    "row_number": row_number,
                    "url": url,
                    "Parser Status": "RETRY",
                    "Worker": worker_id,
                    "Attempts": attempt,
                    "HTML Time": html_time,
                    "Error": str(exc),
                }

            return {
                "row_number": row_number,
                "url": url,
                "Parser Status": "ERROR",
                "Worker": worker_id,
                "Attempts": attempt,
                "HTML Time": html_time,
                "Error": str(exc),
            }

        html_time = (
            time.perf_counter()
            - html_start
        )

        html_size = len(html)

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"HTML: {html_size:,} chars"
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"HTML time: {html_time:.3f}s"
        )

        # =====================================================================
        # CAPTCHA DETECTION
        # =====================================================================

        html_lower = html.lower()

        antibot = (
            "captcha" in html_lower
            or "access denied" in html_lower
            or "ozon-antibot" in html_lower
        )

        if antibot:

            total_time = (
                time.perf_counter()
                - started
            )

            self.logger.log(
                f"[WORKER {worker_id}] "
                f"!!! ANTIBOT DETECTED !!!"
            )

            return {
                "row_number": row_number,
                "url": url,
                "Parser Status": "CAPTCHA",
                "HTTP Status": status,
                "Final URL": final_url,
                "Worker": worker_id,
                "Attempts": attempt,
                "Goto Time": goto_time,
                "HTML Time": html_time,
                "Total Time": total_time,
                "Error":
                    "Ozon antibot detected",
            }

        # =====================================================================
        # PARSER
        # =====================================================================

        parser_start = time.perf_counter()

        try:

            # Парсинг HTML - CPU-bound работа, выполняем в ОТДЕЛЬНОМ
            # ПРОЦЕССЕ (не в этом потоке), чтобы не терять время на
            # GIL-контention с другими browser-воркерами, которые
            # тоже могут парсить HTML параллельно в своих потоках.
            future = self.parser_pool.submit(
                _parse_html_in_process, html
            )
            parsed = future.result(timeout=30)

            parser_error = parsed.pop("__error__", None)

            if parser_error:
                self.logger.log(
                    f"[WORKER {worker_id}] "
                    f"✗ PARSER ERROR: {parser_error}"
                )
                parsed = {}

        except Exception as exc:

            parser_error = str(exc)

            parsed = {}

            self.logger.log(
                f"[WORKER {worker_id}] "
                f"✗ PARSER ERROR: "
                f"{repr(exc)}"
            )

        parser_time = (
            time.perf_counter()
            - parser_start
        )

        # =====================================================================
        # PARSED VALUES
        # =====================================================================

        title = parsed.get(
            "title",
            "",
        )

        description = parsed.get(
            "description",
            "",
        )

        price = parsed.get(
            "price"
        )

        currency = parsed.get(
            "currency"
        )

        sku = parsed.get(
            "sku",
            "",
        )

        brand = parsed.get(
            "brand",
            "",
        )

        main_image = parsed.get(
            "main_image",
            "",
        )

        images = parsed.get(
            "images",
            [],
        )

        if not isinstance(
            images,
            list,
        ):

            images = []

        specs = parsed.get(
            "specs",
            {},
        )

        if not isinstance(
            specs,
            dict,
        ):

            specs = {}

        material = parsed.get(
            "material",
            "",
        )

        # =====================================================================
        # SUCCESS (только парсинг - классификация вынесена в отдельный
        # однопоточный проход classify_all(), см. PHASE 2 в run())
        # =====================================================================

        parser_success = (
            parser_error is None
            and bool(
                title
                or sku
                or price is not None
            )
        )

        total_time = (
            time.perf_counter()
            - started
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"TITLE: {str(title)[:100]}"
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"SKU: {sku}"
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"PRICE: {price}"
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"IMAGES: {len(images)}"
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"PARSER: {parser_time:.3f}s"
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"TOTAL: {total_time:.3f}s"
        )

        self.logger.log(
            f"[WORKER {worker_id}] "
            f"{'✓ OK' if parser_success else '✗ ERROR'}"
        )

        # Копим сырые тайминги для перцентилей в финальном отчёте.
        with self.results_lock:
            self.timings["goto"].append(goto_time)
            self.timings["html"].append(html_time)
            self.timings["parser"].append(parser_time)
            self.timings["total"].append(total_time)
            self.timings["html_size"].append(html_size)

        return {
            "row_number": row_number,
            "url": url,
            "Parser Status":
                "OK"
                if parser_success
                else "ERROR",
            "HTTP Status": status,
            "Final URL": final_url,
            "Title": title,
            "Description": description,
            "Price": price,
            "Currency": currency,
            "SKU": sku,
            "Brand": brand,
            "Main Image": main_image,
            "Images": json.dumps(
                images,
                ensure_ascii=False,
            ),
            "Images Count": len(images),
            "Specs": json.dumps(
                specs,
                ensure_ascii=False,
            ),
            "Specs Count": len(specs),
            "Material": material,
            "Worker": worker_id,
            "Attempts": attempt,
            "Goto Time": round(
                goto_time,
                3,
            ),
            "HTML Time": round(
                html_time,
                3,
            ),
            "Parser Time": round(
                parser_time,
                3,
            ),
            "Total Time": round(
                total_time,
                3,
            ),
            "HTML Size": html_size,
            "Error":
                parser_error or "",
            # Служебные поля для PHASE 2 (classify_all) - не колонки
            # Excel, отфильтровываются автоматически в write_result(),
            # т.к. их нет в RESULT_COLUMNS.
            "_parsed_raw": parsed,
            "_final_url": final_url,
        }

    # =========================================================================
    # WRITER
    # =========================================================================

    def writer(self):

        processed_since_save = 0

        while True:

            try:

                result = (
                    self.result_queue.get(
                        timeout=0.5
                    )
                )

            except queue.Empty:

                if (
                    self.work_queue.unfinished_tasks
                    == 0
                    and all(
                        not t.is_alive()
                        for t in self.worker_threads
                    )
                ):

                    break

                continue

            try:

                status = result.get(
                    "Parser Status"
                )

                # -------------------------------------------------------------
                # RETRY
                # -------------------------------------------------------------

                if status == "RETRY":

                    # RETRY-задание будет выполнено worker'ом.
                    #
                    # Сам результат retry в Excel пока не пишем.
                    #
                    pass

                else:

                    self.write_result(
                        result
                    )

                    # PHASE 2: копим удачно спарсенные карточки для
                    # последующей однопроходной классификации.
                    if status == "OK" and "_parsed_raw" in result:

                        with self.results_lock:
                            self.phase1_results.append({
                                "row_number": result["row_number"],
                                "url": (
                                    result.get("_final_url")
                                    or result.get("url")
                                ),
                                "parsed": result["_parsed_raw"],
                            })

                    with self.results_lock:

                        self.completed += 1

                        if status == "OK":

                            self.success_count += 1

                        elif status == "CAPTCHA":

                            self.captcha_count += 1

                        else:

                            self.error_count += 1

                    processed_since_save += 1

                    self.print_progress(
                        result
                    )

                    if (
                        processed_since_save
                        >= CHECKPOINT_EVERY
                    ):

                        self.save_checkpoint()

                        processed_since_save = 0

            except Exception as exc:

                self.logger.log(
                    "[WRITER] ERROR:"
                )

                self.logger.log(
                    traceback.format_exc()
                )

                self.writer_exception = exc

            finally:

                self.result_queue.task_done()

        # ---------------------------------------------------------------------
        # Final save
        # ---------------------------------------------------------------------

        self.save_checkpoint()

    # =========================================================================
    # WRITE RESULT
    # =========================================================================

    def write_result(
        self,
        result: dict,
    ):

        row_number = result[
            "row_number"
        ]

        for key in RESULT_COLUMNS:

            if key not in result:

                continue

            column = self.headers.get(
                key
            )

            if not column:

                continue

            value = result.get(
                key
            )

            self.ws.cell(
                row=row_number,
                column=column,
                value=value,
            )

    # =========================================================================
    # CHECKPOINT
    # =========================================================================

    def save_checkpoint(self):

        try:

            self.wb.save(
                self.output_file
            )

            self.last_save_time = time.time()

            self.logger.log(
                f"[CHECKPOINT] ✓ "
                f"{self.output_file}"
            )

        except Exception as exc:

            self.logger.log(
                f"[CHECKPOINT] ✗ ERROR: "
                f"{repr(exc)}"
            )

    # =========================================================================
    # PROGRESS
    # =========================================================================

    def print_progress(
        self,
        result,
    ):

        elapsed = (
            time.perf_counter()
            - self.start_time
        )

        speed = (
            self.completed / elapsed
            if elapsed > 0
            else 0
        )

        per_hour = (
            speed * 3600
        )

        per_day = (
            speed * 86400
        )

        remaining = (
            self.total_urls
            - self.completed
        )

        eta = (
            remaining / speed
            if speed > 0
            else 0
        )

        self.logger.log(
            ""
        )

        self.logger.log(
            f"[PROGRESS] "
            f"{self.completed}/"
            f"{self.total_urls} "
            f"| "
            f"{result.get('Parser Status')} "
            f"| "
            f"{result.get('Total Time', 0)}s "
            f"| "
            f"{speed:.3f} URL/s "
            f"| "
            f"{per_hour:,.0f}/hour "
            f"| "
            f"{per_day:,.0f}/day "
            f"| "
            f"ETA {eta / 60:.1f} min"
        )

    # =========================================================================
    # RUN
    # =========================================================================

    def run(self):

        self.start_time = time.perf_counter()

        self.open_excel()

        self.read_urls()

        if self.total_urls == 0:

            self.logger.log(
                ""
            )

            self.logger.log(
                "Нет новых Ozon URL."
            )

            return

        self.logger.log(
            ""
        )

        self.logger.log(
            "=" * 100
        )

        self.logger.log(
            "START PROCESSING"
        )

        self.logger.log(
            "=" * 100
        )

        self.logger.log(
            f"URLs:    {self.total_urls}"
        )

        self.logger.log(
            f"Workers: {self.workers}"
        )

        self.logger.log(
            f"CDP:     {CDP_URL}"
        )

        self.logger.log(
            "=" * 100
        )

        # =====================================================================
        # START WORKERS
        # =====================================================================

        self.worker_threads = []

        for worker_id in range(
            1,
            self.workers + 1,
        ):

            thread = threading.Thread(
                target=self.worker,
                args=(worker_id,),
                name=(
                    f"OzonWorker-{worker_id}"
                ),
                daemon=True,
            )

            self.worker_threads.append(
                thread
            )

            thread.start()

        # =====================================================================
        # START WRITER
        # =====================================================================

        self.writer_thread = (
            threading.Thread(
                target=self.writer,
                name="ExcelWriter",
                daemon=True,
            )
        )

        self.writer_thread.start()

        # =====================================================================
        # WAIT WORKERS
        # =====================================================================

        for thread in self.worker_threads:

            thread.join()

        # =====================================================================
        # WAIT QUEUES
        # =====================================================================

        self.work_queue.join()
        self.result_queue.join()
        self.writer_thread.join()

        # =====================================================================
        # SHUTDOWN PARSER POOL
        # =====================================================================

        self.parser_pool.shutdown(wait=True)

        self.phase1_elapsed = (
            time.perf_counter() - self.start_time
        )

        self.logger.log("")
        self.logger.log("=" * 100)
        self.logger.log(
            f"PHASE 1 (парсинг) ЗАВЕРШЕНА: "
            f"{self.phase1_elapsed:.2f} сек, "
            f"{len(self.phase1_results)} карточек готовы "
            f"к классификации"
        )
        self.logger.log("=" * 100)

        # =====================================================================
        # PHASE 2: CLASSIFY (единый последовательный проход)
        # =====================================================================

        self.classify_all()

        # =====================================================================
        # FINAL CARD REPOSITORY FLUSH
        #
        # decide() теперь пишет на диск не на каждый вызов, а раз в
        # N (_flush_every) - гарантируем здесь, что последние
        # накопленные в памяти карточки тоже попадут в файл.
        # =====================================================================

        try:
            self.decision_engine.knowledge.card_repository.flush()
        except Exception:
            self.logger.log(
                "[FINAL FLUSH] ✗ ERROR:"
            )
            self.logger.log(
                traceback.format_exc()
            )

        # =====================================================================
        # FINAL SAVE
        # =====================================================================

        self.save_checkpoint()
        self.print_summary()
        self.print_performance_profile()

    # =========================================================================
    # PHASE 2: CLASSIFY_ALL
    #
    # Единый последовательный проход по уже спарсенным карточкам -
    # ОДИН поток, без блокировок, без конкуренции с браузерными
    # воркерами. Если после этого Classify всё ещё медленный - это
    # железно доказывает, что дело в самой логике классификаторов
    # (HistoryClassifier/CardClassifier/ProductResolver и т.п.), а не
    # в многопоточности/локах/flush().
    # =========================================================================

    def classify_all(self):

        total = len(self.phase1_results)

        self.logger.log(
            f"PHASE 2 CLASSIFY START: {total}"
        )

        start = time.perf_counter()

        futures = []

        workers = CLASSIFIER_PROCESSES

        with ProcessPoolExecutor(
                max_workers=workers
        ) as executor:

            for item in self.phase1_results:
                future = executor.submit(
                    process_classifier_task,
                    item,
                    self.decision_engine,
                    build_product_card,
                )

                futures.append(future)

            completed = 0

            for future in as_completed(futures):

                result = future.result()

                completed += 1

                self.write_result({

                    "row_number":
                        result["row_number"],

                    "Product":
                        result["product"],

                    "Display Name":
                        result["display_name"],

                    "TNVED Code":
                        result["code"],

                    "Confidence":
                        result["confidence"],

                    "Source":
                        result["source"],

                    "Review":
                        "Да"
                        if result["review"]
                        else "",

                    "Error":
                        result["error"],

                })

                if completed % 10 == 0:
                    elapsed = (
                            time.perf_counter()
                            - start
                    )

                    speed = (
                            completed / elapsed
                    )

                    self.logger.log(
                        f"CLASSIFY "
                        f"{completed}/{total} "
                        f"{speed:.3f} card/s"
                    )

        self.phase2_elapsed = (
                time.perf_counter()
                - start
        )

        self.logger.log(
            f"PHASE 2 FINISHED "
            f"{self.phase2_elapsed:.2f}s"
        )


    # =========================================================================
    # SUMMARY
    # =========================================================================

    def print_summary(self):

        elapsed = (
            time.perf_counter()
            - self.start_time
        )

        speed = (
            self.completed / elapsed
            if elapsed > 0
            else 0
        )

        per_hour = (
            speed * 3600
        )

        per_day = (
            speed * 86400
        )

        self.logger.log(
            ""
        )

        self.logger.log(
            ""
        )

        self.logger.log(
            "=" * 100
        )

        self.logger.log(
            "ИТОГ"
        )

        self.logger.log(
            "=" * 100
        )

        self.logger.log(
            f"Всего URL:             "
            f"{self.total_urls}"
        )

        self.logger.log(
            f"Обработано:            "
            f"{self.completed}"
        )

        self.logger.log(
            f"Успешно:               "
            f"{self.success_count}"
        )

        self.logger.log(
            f"Ошибки:                "
            f"{self.error_count}"
        )

        self.logger.log(
            f"CAPTCHA:               "
            f"{self.captcha_count}"
        )

        self.logger.log(
            ""
        )

        self.logger.log(
            f"Время:                 "
            f"{elapsed:.2f} сек"
        )

        self.logger.log(
            f"Средняя скорость:      "
            f"{speed:.3f} URL/сек"
        )

        self.logger.log(
            f"В час:                 "
            f"{per_hour:,.0f}"
        )

        self.logger.log(
            f"За 24 часа:            "
            f"{per_day:,.0f}"
        )

        self.logger.log(
            ""
        )

        self.logger.log(
            f"OK:                    "
            f"{self.success_count}"
        )

        self.logger.log(
            f"CAPTCHA:               "
            f"{self.captcha_count}"
        )

        self.logger.log(
            f"ERROR:                 "
            f"{self.error_count}"
        )

        self.logger.log(
            ""
        )

        self.logger.log(
            f"RESULT:                "
            f"{self.output_file}"
        )

        self.logger.log(
            f"LOG:                   "
            f"{LOG_FILE}"
        )

        self.logger.log(
            "=" * 100
        )

    # =========================================================================
    # PERFORMANCE PROFILE (расширенный - перцентили + блокировка ресурсов)
    # =========================================================================

    def _percentile(self, values, pct):

        if not values:
            return 0.0

        sorted_values = sorted(values)
        index = int(len(sorted_values) * pct)
        index = min(index, len(sorted_values) - 1)

        return sorted_values[index]

    def print_performance_profile(self):

        goto = self.timings["goto"]
        html = self.timings["html"]
        parser = self.timings["parser"]
        classify = self.timings["classify"]
        total = self.timings["total"]
        sizes = self.timings["html_size"]

        n = len(total)

        self.logger.log("")
        self.logger.log("=" * 100)
        self.logger.log("PERFORMANCE PROFILE (EXTENDED)")
        self.logger.log("=" * 100)

        # -----------------------------------------------------------------
        # PHASE 1 vs PHASE 2 - раздельное общее время. Это отвечает на
        # главный вопрос: сколько реально стоит парсинг (можно
        # распараллелить браузерами) против классификации (сейчас
        # строго последовательно, один поток, без блокировок).
        # -----------------------------------------------------------------
        phase1 = getattr(self, "phase1_elapsed", 0)
        phase2 = getattr(self, "phase2_elapsed", 0)
        phases_total = phase1 + phase2

        self.logger.log("PHASE 1 (парсинг, многопоточно) vs PHASE 2 (классификация, 1 поток)")
        self.logger.log("-" * 100)
        if phases_total > 0:
            self.logger.log(
                f"Phase 1 - парсинг:       {phase1:8.2f}s "
                f"({phase1 / phases_total * 100:5.1f}%)"
            )
            self.logger.log(
                f"Phase 2 - классификация: {phase2:8.2f}s "
                f"({phase2 / phases_total * 100:5.1f}%)"
            )
        self.logger.log("")

        if n == 0:
            self.logger.log("Нет успешных запросов для профиля.")
            self.logger.log("=" * 100)
            return

        def stage_line(name, values):
            avg = sum(values) / len(values) if values else 0
            p50 = self._percentile(values, 0.50)
            p90 = self._percentile(values, 0.90)
            p99 = self._percentile(values, 0.99)
            worst = max(values) if values else 0

            self.logger.log(
                f"{name:<14} avg={avg:6.3f}s  "
                f"p50={p50:6.3f}s  p90={p90:6.3f}s  "
                f"p99={p99:6.3f}s  max={worst:6.3f}s"
            )

        self.logger.log(
            "PERCENTILES ПО ЭТАПАМ "
            "(среднее прячет редкие тяжёлые страницы - "
            "смотрите на p90/p99/max, чтобы найти их)"
        )
        self.logger.log("-" * 100)
        stage_line("Browser goto:", goto)
        stage_line("HTML content:", html)
        stage_line("Parser:", parser)
        stage_line("Classify:", classify)
        stage_line("Total:", total)

        self.logger.log("")
        self.logger.log("РАЗМЕР HTML")
        self.logger.log("-" * 100)

        avg_size = sum(sizes) / len(sizes) if sizes else 0
        max_size = max(sizes) if sizes else 0
        min_size = min(sizes) if sizes else 0

        self.logger.log(
            f"Средний размер:  {avg_size:,.0f} chars"
        )
        self.logger.log(
            f"Мин / Макс:      {min_size:,} / {max_size:,} chars"
        )

        # Корреляция размера HTML со временем парсинга - если она
        # высокая, значит парсер тормозит именно на больших страницах
        # (частый признак: BeautifulSoup + html.parser на документах
        # с большим количеством виджетов рекомендаций/отзывов).
        if len(sizes) > 1 and len(parser) == len(sizes):
            correlation = self._correlation(sizes, parser)
            self.logger.log(
                f"Корреляция размер↔время парсинга: {correlation:.2f} "
                f"(1.0 = чем больше страница, тем дольше парсинг)"
            )

        self.logger.log("")
        self.logger.log("СЕТЕВЫЕ ЗАПРОСЫ")
        self.logger.log("-" * 100)

        total_requests = (
            self.total_blocked_requests
            + self.total_allowed_requests
        )
        blocked_pct = (
            (self.total_blocked_requests / total_requests * 100)
            if total_requests else 0
        )

        self.logger.log(
            f"Заблокировано:   {self.total_blocked_requests:,} "
            f"({blocked_pct:.1f}% от всех запросов)"
        )
        self.logger.log(
            f"Пропущено:       {self.total_allowed_requests:,}"
        )
        self.logger.log(
            f"В среднем на URL: "
            f"{self.total_blocked_requests / n:.1f} заблокировано, "
            f"{self.total_allowed_requests / n:.1f} пропущено"
        )

        self.logger.log("")
        self.logger.log("BOTTLENECK (по сумме времени)")
        self.logger.log("-" * 100)

        total_goto = sum(goto)
        total_html = sum(html)
        total_parser = sum(parser)
        total_classify = sum(classify)
        grand_total = total_goto + total_html + total_parser + total_classify

        if grand_total > 0:
            self.logger.log(
                f"Browser goto:  {total_goto:8.2f}s "
                f"({total_goto / grand_total * 100:5.1f}%)"
            )
            self.logger.log(
                f"HTML content:  {total_html:8.2f}s "
                f"({total_html / grand_total * 100:5.1f}%)"
            )
            self.logger.log(
                f"Parser:        {total_parser:8.2f}s "
                f"({total_parser / grand_total * 100:5.1f}%)"
            )
            self.logger.log(
                f"Classify:      {total_classify:8.2f}s "
                f"({total_classify / grand_total * 100:5.1f}%)"
            )

        self.logger.log("=" * 100)

    def _correlation(self, x, y):
        """Коэффициент корреляции Пирсона, без numpy/scipy."""

        n = len(x)

        if n < 2:
            return 0.0

        mean_x = sum(x) / n
        mean_y = sum(y) / n

        cov = sum(
            (x[i] - mean_x) * (y[i] - mean_y)
            for i in range(n)
        )

        var_x = sum((v - mean_x) ** 2 for v in x)
        var_y = sum((v - mean_y) ** 2 for v in y)

        denom = (var_x * var_y) ** 0.5

        if denom == 0:
            return 0.0

        return cov / denom


# =============================================================================
# MAIN
# =============================================================================


def main():

    # ОБЯЗАТЕЛЬНО для ProcessPoolExecutor/multiprocessing в приложении,
    # собранном через PyInstaller (--onefile) на Windows. Без этого
    # каждый дочерний процесс пытается заново запустить весь .exe -
    # бесконечное размножение процессов.
    import multiprocessing
    multiprocessing.freeze_support()

    print()
    print("=" * 100)
    print("OZON EXCEL PROCESSOR")
    print("=" * 100)

    input_file = DEFAULT_INPUT_FILE

    output_file = None

    processor = OzonExcelProcessor(
        input_file=input_file,
        output_file=output_file,
        workers=2,
    )

    processor.run()


if __name__ == "__main__":

    main()