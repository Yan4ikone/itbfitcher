import threading
import traceback
import time
import random
import asyncio
import openpyxl

from learning.importer import load_learning_history
from parser.cdp_product_parser import CDPProductParser, BLOCKED_RESOURCE_TYPES, log
from engines.decision_engine import DecisionEngine
from modules.decision_logger import DecisionLogger
from pathlib import Path
from repositories.card_repository import CardRepository


class OzonAutoProcessor:

    # ==========================================================
    # НАСТРОЙКА МНОГОПОТОЧНОСТИ
    #
    # Снижено с 4 до 2: на 4 воркерах результат стабильно хуже
    # однопоточного (11/37 вместо ожидаемого улучшения) - похоже на
    # антибот/перегрузку браузера при параллельной отрисовке тяжёлых
    # SPA-страниц. Начинайте с 2 и поднимайте постепенно, только
    # если на реальном файле результат стабильно не хуже, чем при
    # меньшем значении.
    # ==========================================================
    MAX_WORKERS = 2
    # Пауза между запросами ОДНОГО воркера. Без неё 4 вкладки бьют по
    # Ozon без остановки, и после ~8-15 быстрых запросов подряд
    # срабатывает антибот-защита - страница отдаётся урезанной, без
    # нужных виджетов, и парсинг падает по таймауту сразу на всех
    # товарах. Раньше эта пауза (random 2.0-4.0с) была в однопоточном
    # коде между КАЖДЫМ товаром; теперь она на уровне каждого воркера -
    # общая скорость всё равно кратно выше, потому что 4 воркера ждут
    # параллельно, а не по очереди.
    WORKER_DELAY_MIN = 1.5
    WORKER_DELAY_MAX = 3.0

    def __init__(
            self,
            excel_path,
            logger=None,
            stats_callback=None,
            limit=None,
            skip_filled=False
    ):
        self.skip_filled = skip_filled
        self.excel_path = excel_path
        self.logger = logger
        self.limit = limit
        self.stats_callback = stats_callback
        self.stop_requested = False
        self.pause_requested = False
        self.pause_event = threading.Event()
        self.pause_event.set()
        self.total_rows = 0
        self.processed_rows = 0
        self.found_count = 0
        self.not_found_count = 0
        self.cached_count = 0
        self.learning_buffer = []
        p = Path(excel_path)
        self.result_path = str(
            p.with_name(
                f"{p.stem}_RESULT{p.suffix}"
            )
        )
        self.decision_logger = DecisionLogger()
        self.card_repository = CardRepository()
        # ------------------------------------------------------
        # Защита общей статистики
        # ------------------------------------------------------
        self.stats_lock = threading.Lock()
        # ------------------------------------------------------
        # DecisionEngine нельзя одновременно заставлять
        # писать CardRepository из нескольких потоков.
        # Поэтому decide() будет защищён этим lock.
        # ------------------------------------------------------
        self.engine_lock = threading.Lock()
    # ==========================================================
    # LOG
    # ==========================================================
    def log(self, text):

        print(text)

        if self.logger:
            self.logger(text)
    # ==========================================================
    # CACHE
    # ==========================================================
    def get_cached_card(self, url):

        if not url:
            return None

        card = self.card_repository.find_by_url(url)

        if card:
            return card

        return self.card_repository.find_by_normalized_url(url)
    # ==========================================================
    # PAUSE / RESUME / STOP
    # ==========================================================
    def pause(self):

        self.pause_requested = True
        self.pause_event.clear()
        self.log("Обработка приостановлена")

    def resume(self):

        self.pause_requested = False
        self.pause_event.set()
        self.log("Обработка продолжена")

    def stop(self):

        self.stop_requested = True

        # Разбудить ожидающие worker,
        # чтобы они могли увидеть stop_requested.
        self.pause_event.set()
        self.log("Получена команда остановки...")

    async def async_worker(self, worker_id, page, task_queue, result_queue, engine):

        self.log(
            f"[OzonWorker-{worker_id}] "
            f"запущен"
        )
        first_task = True
        # ==================================================
        # СТАРТОВЫЙ РАЗБРОС
        #
        # Раньше пауза пропускалась перед ПЕРВЫМ запросом каждого
        # воркера ("if not first_task") - из-за этого все N воркеров
        # стартовали и били по Ozon ОДНОВРЕМЕННО в первую же секунду,
        # что и приводило к пачке из 2-4 вкладок, открывающихся разом,
        # и повышенному риску капчи. Теперь у каждого воркера свой
        # стартовый сдвиг, пропорциональный его номеру - они расходятся
        # во времени с самого начала, а не только начиная со второго
        # запроса.
        # ==================================================
        startup_delay = worker_id * random.uniform(1.0, 2.0)
        await asyncio.sleep(startup_delay)

        while True:

            task = await task_queue.get()

            try:
                if task is None:
                    return

                row, url, excel_name = task
                # ==================================================
                # STOP
                # ==================================================
                if self.stop_requested:
                    await result_queue.put({
                        "row": row,
                        "url": url,
                        "stopped": True,
                    })
                    continue
                # ==================================================
                # PAUSE
                # ==================================================
                while self.pause_requested:
                    await asyncio.sleep(
                        0.2
                    )
                if self.stop_requested:
                    await result_queue.put({
                        "row": row,
                        "url": url,
                        "stopped": True,
                    })
                    continue
                # ==================================================
                # CACHE
                # ==================================================
                cached_card = (self.get_cached_card(url))

                if cached_card:
                    await result_queue.put({
                        "row": row,
                        "url": url,
                        "cached": True,
                        "cached_card": cached_card,
                    })
                    continue
                # ==================================================
                # АНТИБОТ-ПАУЗА (между запросами ОДНОГО воркера,
                # включая первый - стартовый разброс выше решает
                # только проблему одновременного старта ВСЕХ воркеров,
                # а эта пауза нужна для каждого следующего запроса
                # того же воркера).
                # ==================================================
                if not first_task:
                    delay = random.uniform(
                        self.WORKER_DELAY_MIN,
                        self.WORKER_DELAY_MAX,
                    )
                    await asyncio.sleep(delay)
                first_task = False
                # ==================================================
                # LOG
                # ==================================================
                self.log(
                    f"[OzonWorker-{worker_id}] "
                    f"Строка {row}: {url}"
                )
                # ==================================================
                # PARSE
                # ==================================================
                card = await self.async_parser.parse_url_async(page, url)
                # ==================================================
                # EXCEL TITLE
                # ==================================================
                if excel_name:
                    card.excel_title = excel_name
                # ==================================================
                # RESULT
                # ==================================================
                await result_queue.put({
                    "row": row,
                    "url": url,
                    "cached": False,
                    "card": card,
                })
            except Exception as exc:
                await result_queue.put({
                    "row": task[0],
                    "url": task[1],
                    "error": exc,
                    "traceback": traceback.format_exc(),
                })
            finally:

                task_queue.task_done()

    async def _run_async(self, rows_to_process, ws, wb, engine):
        """
        Главный async pipeline.
        Один Playwright connection.
        Один Browser Context.
        Четыре постоянные страницы.
        """
        self.async_parser = (CDPProductParser())
        # ======================================================
        # ONE CDP CONNECTION
        # ======================================================
        await self.async_parser.connect_async()

        context = (self.async_parser.async_context)

        self.log(
            "CDP подключён. "
            "Создаём постоянные вкладки..."
        )
        # ======================================================
        # 4 ПОСТОЯННЫЕ ВКЛАДКИ
        # ======================================================
        worker_count = min(self.MAX_WORKERS, len(rows_to_process))
        # ======================================================
        # ROUTE HANDLER (единственное определение - используется
        # для ВСЕХ вкладок, и существующих, и новых. Раньше был
        # дублирующийся локальный route_handler внутри цикла добора
        # вкладок, из-за чего на новых вкладках регистрировалось
        # ДВА обработчика маршрутов на один и тот же паттерн "**/*" -
        # лишние накладные расходы на каждый сетевой запрос.)
        # ======================================================
        async def route_handler(route):

            try:
                if (
                        route.request.resource_type
                        in BLOCKED_RESOURCE_TYPES
                ):
                    await route.abort()

                else:
                    await route.continue_()
            except Exception:
                try:
                    await route.continue_()
                except Exception:
                    pass
        # ======================================================
        # ПОЛУЧАЕМ УЖЕ СУЩЕСТВУЮЩИЕ ВКЛАДКИ
        # ======================================================
        pages = [
            page
            for page in context.pages
            if not page.is_closed()
        ]
        self.log(f"Существующих вкладок: {len(pages)}")
        # ======================================================
        # ЕСЛИ НЕТ ВКЛАДОК — СОЗДАЁМ ПЕРВУЮ
        # ======================================================
        if not pages:

            page = (await self.async_parser.async_browser.new_page())
            pages.append(page)
        # ======================================================
        # ДОБИРАЕМ ВКЛАДКИ ДО worker_count
        # ======================================================
        while len(pages) < worker_count:

            page = await context.new_page()
            pages.append(page)

            self.log(
                f"[OzonWorker-{len(pages)}] "
                f"Page готова: {page.url}"
            )
            # --------------------------------------------------
            # Пауза между созданием вкладок
            # --------------------------------------------------
            if len(pages) < worker_count:
                await asyncio.sleep(0.5)
        # ======================================================
        # НАСТРАИВАЕМ ВСЕ РАБОЧИЕ ВКЛАДКИ (ровно один route на
        # каждую, включая уже существующие)
        # ======================================================
        for index, page in enumerate(
                pages[:worker_count],
                start=1,
        ):
            await page.route(
                "**/*",
                route_handler,
            )
            self.log(
                f"[OzonWorker-{index}] "
                f"Page готова: {page.url}"
            )
        # ======================================================
        # ЕСЛИ БРАУЗЕР БЫЛ ЗАПУЩЕН С ЛИШНИМИ ВКЛАДКАМИ
        # ======================================================
        if len(pages) > worker_count:

            for page in pages[worker_count:]:
                try:
                    await page.close()
                except Exception:
                    pass

            pages = pages[:worker_count]
        self.log(f"Рабочих вкладок: {len(pages)}")
        # ======================================================
        # QUEUES
        # ======================================================
        task_queue = asyncio.Queue()
        result_queue = asyncio.Queue()
        # ======================================================
        # ЗАДАЧИ
        # ======================================================
        for task in rows_to_process:
            if self.stop_requested:
                break
            await task_queue.put(
                task
            )
        # ======================================================
        # WORKERS
        # ======================================================
        workers = []

        for index, page in enumerate(
                pages
        ):
            workers.append(
                asyncio.create_task(
                    self.async_worker(
                        index + 1,
                        page,
                        task_queue,
                        result_queue,
                        engine,
                    )
                )
            )
        # ======================================================
        # STOP SIGNAL
        # ======================================================
        for _ in workers:
            await task_queue.put(
                None
            )
        # ======================================================
        # RESULTS
        # ======================================================
        completed = 0

        try:
            while (
                    completed
                    < len(rows_to_process)
            ):
                result_data = (
                    await result_queue.get()
                )
                completed += 1

                try:
                    # ==============================================
                    # ERROR
                    # ==============================================
                    if "error" in result_data:
                        self.log(
                            result_data.get(
                                "traceback",
                                "Неизвестная ошибка",
                            )
                        )
                        self.processed_rows += 1
                        self.print_progress()

                        continue
                    # ==============================================
                    # STOP
                    # ==============================================
                    if result_data.get("stopped"):
                        self.processed_rows += 1
                        self.print_progress()

                        continue

                    row = result_data["row"]
                    # ==============================================
                    # CACHE
                    # ==============================================
                    if result_data.get("cached"):

                        self.apply_cached_result(
                            ws,
                            row,
                            result_data[
                                "cached_card"
                            ],
                        )
                        self.cached_count += 1
                    # ==============================================
                    # NORMAL CARD
                    # ==============================================
                    else:

                        card = result_data["card"]
                        # ------------------------------------------
                        # DecisionEngine теперь выполняется здесь
                        # ------------------------------------------
                        with self.engine_lock:

                            result = engine.decide(card)
                        self.apply_result(
                            ws,
                            row,
                            card,
                            result,
                        )
                    # ==============================================
                    # PROGRESS
                    # ==============================================
                    self.processed_rows += 1
                    self.print_progress()
                    # ==============================================
                    # SAVE
                    # ==============================================
                    if (
                            self.processed_rows
                            % 20
                            == 0
                    ):
                        wb.save(self.result_path)
                        self.log("Файл сохранён")

                finally:
                    result_queue.task_done()
        finally:
            # ==================================================
            # WAIT WORKERS
            # ==================================================
            await asyncio.gather(
                *workers,
                return_exceptions=True,
            )
            # ==================================================
            # CLOSE 4 PAGES
            # ==================================================
            for page in pages:
                try:
                    await page.close()
                except Exception:

                    log.debug(
                        "Ошибка закрытия страницы",
                        exc_info=True,
                    )
            # ==================================================
            # CLOSE PLAYWRIGHT
            # ==================================================
            await self.async_parser.disconnect_async(close_browser=False)
    # ==========================================================
    # URL
    # ==========================================================

    def get_url_from_row(self, ws, row):

        for col in ("D", "E"):

            cell = ws[f"{col}{row}"]

            if cell.hyperlink:
                return cell.hyperlink.target

            value = cell.value

            if (
                    isinstance(value, str)
                    and value.startswith("http")
            ):
                return value
        return None
    # ==========================================================
    # APPLY CACHE RESULT
    # ==========================================================

    def apply_cached_result(
            self,
            ws,
            row,
            cached_card,
    ):
        description = (
            cached_card.get("display_name")
            or cached_card.get("product")
            or cached_card.get("description")
            or ""
        )
        code = str(
            cached_card.get(
                "code",
                "",
            )
        ).strip()

        if description:
            ws[f"B{row}"] = description

        if code and code not in (
                "0",
                "nan",
        ):
            try:
                ws[f"C{row}"] = int(code)
            except ValueError:
                ws[f"C{row}"] = code

            self.found_count += 1
            self.log(f"CACHE Описание: {description}")
            self.log(f"CACHE Код: {code}")

        else:

            self.not_found_count += 1

    # ==========================================================
    # APPLY NORMAL RESULT
    # ==========================================================

    def apply_result(self, ws, row, card, result):
        # ВАЖНО:
        # card.excel_title должен быть установлен ДО decide().
        #
        # Это будет сделано в run(), перед отправкой worker.
        # Если decide() уже был вызван без него, значение B
        # не влияет на уже полученный результат.
        #
        # Поэтому ниже ничего не меняем.
        # ------------------------------------------------------
        ws[f"K{row}"] = (
            "Да"
            if result.new_product
            else ""
        )
        ws[f"L{row}"] = (
            "Да"
            if result.new_dropdown
            else ""
        )
        # ------------------------------------------------------
        # Decision Logger
        # ------------------------------------------------------
        self.decision_logger.save(card, result)
        # ------------------------------------------------------
        # Learning buffer
        # ------------------------------------------------------
        self.learning_buffer.append(
            {
                "url": card.url,
                "title": card.title,
                "description": card.description,
                "product": result.product,
                "code": result.code,
                "material": result.material,
            }
        )
        # ------------------------------------------------------
        # B
        # ------------------------------------------------------
        original = ws[f"B{row}"].value

        if result.dropdown:
            ws[f"B{row}"] = (result.dropdown)
        elif result.product:
            ws[f"B{row}"] = (result.product)
        else:
            ws[f"B{row}"] = original
        # ------------------------------------------------------
        # C
        # ------------------------------------------------------
        if result.code:
            try:
                ws[f"C{row}"] = int(result.code)
            except ValueError:
                ws[f"C{row}"] = (result.code)
        # ------------------------------------------------------
        # Statistics
        # ------------------------------------------------------

        if result.code:

            self.found_count += 1
            self.log(f"Описание: {result.dropdown}")
            self.log(f"Товар: {result.product}")
            self.log(f"Материал: {result.material}")
            self.log(f"Код: {result.code}")
            self.log(f"Источник: {result.source}")
            self.log(
                f"Уверенность: "
                f"{result.confidence}%"
            )
            if result.review:
                self.log(
                    "⚠ Требуется проверка"
                )
        else:

            self.not_found_count += 1
            self.log(
                f"Описание: "
                f"{result.product}"
            )
            self.log("Код не найден")
    # ==========================================================
    # RUN
    # ==========================================================
    def run(self):

        self.start_time = time.time()
        self.log("Открытие Excel...")
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        rows_to_process = []
        # ==========================================================
        # СОБИРАЕМ СТРОКИ
        # ==========================================================
        for row in range(2, ws.max_row + 1):

            url = self.get_url_from_row(ws, row)
            current_code = ws[f"C{row}"].value

            if self.skip_filled:

                if current_code not in (None, "", 0):
                    continue

            if url:
                excel_name = str(ws[f"B{row}"].value or "").strip()
                rows_to_process.append((row, url, excel_name))
        # ==========================================================
        # LIMIT
        # ==========================================================
        if self.limit:
            rows_to_process = (rows_to_process[ :self.limit])
        self.total_rows = len(rows_to_process)
        self.log(
            f"Всего строк: "
            f"{self.total_rows}"
        )

        if not rows_to_process:

            wb.save(self.result_path)
            self.print_summary()

            return
        # ==========================================================
        # LEARNING HISTORY
        # ==========================================================
        learning_history = (load_learning_history(self.excel_path))
        engine = DecisionEngine(learning_history)
        # ==========================================================
        # ASYNC PARSING
        # ==========================================================
        try:

            asyncio.run(self._run_async(rows_to_process, ws, wb, engine))

        except Exception:

            self.log(traceback.format_exc())

            raise
        finally:

            wb.save(self.result_path)
        self.print_summary()
    # ==========================================================
    # PROGRESS
    # ==========================================================
    def print_progress(self):

        remaining = (
            self.total_rows
            - self.processed_rows
        )
        self.log(
            (
                f"Обработано: "
                f"{self.processed_rows}/"
                f"{self.total_rows} | "
                f"Осталось: "
                f"{remaining}"
            )
        )
        if self.stats_callback:

            self.stats_callback(
                self.total_rows,
                self.processed_rows,
                self.found_count,
                self.not_found_count,
            )
    # ==========================================================
    # SUMMARY
    # ==========================================================
    def print_summary(self):

        self.log("\n")
        self.log("=" * 60)
        self.log("ОБРАБОТКА ЗАВЕРШЕНА")
        self.log(f"Всего: {self.total_rows}")
        self.log(f"Найдено: {self.found_count}")
        self.log(
            f"Не найдено: "
            f"{self.not_found_count}"
        )
        self.log(
            f"Из кеша: "
            f"{self.cached_count}"
        )
        if self.total_rows:

            percent = round(
                (
                    self.found_count
                    / self.total_rows
                ) * 100,
                2,
            )
            self.log(
                f"Успешность: "
                f"{percent}%"
            )
        self.log("=" * 60)
        elapsed = round(
            time.time()
            - self.start_time
        )
        self.log(
            f"Время работы: "
            f"{elapsed} сек."
        )
        if self.processed_rows:

            avg = (
                elapsed
                / self.processed_rows
            )
            self.log(
                f"Среднее на товар: "
                f"{avg:.2f} сек."
            )
        if self.stats_callback:

            self.stats_callback(
                self.total_rows,
                self.total_rows,
                self.found_count,
                self.not_found_count,
            )