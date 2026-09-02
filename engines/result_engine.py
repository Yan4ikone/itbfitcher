from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from dictionaries.all_dictionaries import MATERIAL_COLORS



def apply_result(ws, row, code_col, result):
    set_code(ws, row, code_col, result)
    set_comment(ws, row, code_col, result)
    set_color(ws, row, code_col, result)


def set_code(ws, row, code_col, result):
    cell = ws.cell(row=row, column=code_col)
    try:
        cell.value = int(result.code)
    except (ValueError, TypeError):
        cell.value = result.code


def set_comment(ws, row, code_col, result):
    if not result.review:
        return

    alternatives = result.alternatives or {}

    text = "\n".join(
        f"{code} — {name}"
        for code, name in alternatives.items()
    )

    ws.cell(row=row, column=code_col).comment = Comment(text, "Classifier")

def _get_color_group(result):
    """
    Определяет визуальную группу для результата.

    MATERIAL_COLORS используется как единый справочник цветов
    для материалов, пола, возраста и групп.

    Приоритет:
        1. material_group
        2. material
        3. dropdown_group
        4. matched_features
        5. color
    """
    # ------------------------------------------------------------
    # 1. Явно определённая группа материала
    # ------------------------------------------------------------
    values = [
        getattr(result, "material_group", ""),
        getattr(result, "material", ""),
        getattr(result, "dropdown_group", ""),
    ]
    # ------------------------------------------------------------
    # 2. Признаки, полученные Decision Engine
    # ------------------------------------------------------------
    matched_features = getattr(
        result,
        "matched_features",
        {},
    )

    if isinstance(matched_features, dict):
        # Проверяем сначала наиболее важные признаки.
        for key in (
            "material",
            "material_group",
            "gender",
            "sex",
            "пол",
            "age",
            "возраст",
            "group",
            "product_group",
            "tool_group",
            "type",
            "тип",
        ):
            value = matched_features.get(key)

            if value:
                values.append(value)
    # ------------------------------------------------------------
    # 3. Дополнительный цвет, если он уже был определён
    # ------------------------------------------------------------
    color = getattr(result, "color", "")

    if color:
        values.append(color)
    # ------------------------------------------------------------
    # 4. Ищем соответствие в MATERIAL_COLORS
    # ------------------------------------------------------------
    for value in values:

        if value is None:
            continue

        # Иногда признак может прийти списком
        if isinstance(value, (list, tuple, set)):

            for item in value:

                if not item:
                    continue

                normalized = str(
                    item
                ).strip().lower()

                if normalized in MATERIAL_COLORS:
                    return normalized

                for group in MATERIAL_COLORS:

                    if group in normalized:
                        return group

            continue

        # Иногда признак может прийти словарём
        if isinstance(value, dict):

            for item in value.values():

                if not item:
                    continue

                normalized = str(
                    item
                ).strip().lower()

                if normalized in MATERIAL_COLORS:
                    return normalized

                for group in MATERIAL_COLORS:

                    if group in normalized:
                        return group

            continue

        normalized = str(
            value
        ).strip().lower()

        if not normalized:
            continue

        # Точное совпадение
        if normalized in MATERIAL_COLORS:
            return normalized

        # Частичное совпадение
        for group in MATERIAL_COLORS:

            if group in normalized:
                return group

    return None


def set_color(ws, row, code_col, result):
    """
    Красит ячейку кода согласно определённому признаку.
    ВАЖНО:
        Эта функция НЕ красит всю строку.
        Красная/зелёная строка является ответственностью
        excel.postprocessing.py.
    """
    group = _get_color_group(result)

    if not group:
        return

    color = MATERIAL_COLORS.get(group)

    if not color:
        return

    ws.cell(
        row=row,
        column=code_col,
    ).fill = PatternFill(
        fill_type="solid",
        fgColor=color,
    )