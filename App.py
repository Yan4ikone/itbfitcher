import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import os
import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# --- 1. ЗАПРЕЩЁННЫЕ КОДЫ (из вашего первого макроса) ---
# Эти коды нужно закрашивать красным в Excel. Мы будем их просто помечать.
FORBIDDEN_CODES = [
    "8403", "8407", "8419", "9023", "8426", "8427", "8428", "8429",
    "8430", "8433", "8434", "8435", "8436", "8437", "8438", "8439",
    "8440", "8441", "8442", "8443", "8444", "8447", "8448", "8449",
    "8453", "8454", "8455", "8456", "8457", "8458", "8459", "8460",
    "8461", "8462", "8463", "8464", "8465", "8466", "8467", "8468",
    "8471", "8472", "8473", "8474", "8475", "8476", "8477", "8478",
    "8479", "8480", "8505", "8510", "8511", "8512", "8513", "8517",

]

# --- 2. ЗАПРЕЩЁННЫЕ СЛОВА И СИНОНИМЫ (из вашего второго макроса) ---
# Если слово из левого столбца найдено в наименовании, оно заменяется на слово из правого.
FORBIDDEN_WORDS = {
    r"\bоплата\b": "",
    r"\bпатрон\b": "",
    r"\bювелир.*\b": "",
    r"\bсеребр.*\b": "",
    r"\bзолот.*\b": "",
    r"\bжемчу.*\b": "",
    r"\bохот.*\b": "",
    r"\bпистол.*\b": "",
    r"\bвинтовк.*\b": "",
    r"\bпейнтбол.*\b": "",
    r"\bстрайкбол.*\b": "",
    r"\bстоматолог.*\b": "",
    r"\bприцел.*\b": "",
    r"\bкоптер.*\b|дрон.*\b": "",
    r"\bигла.*\b|игол.*\b|иглы\b": "",
    r"\bалмаз.*\b": "",
    r"\bсекс.*\b": "",
    r"\bдропшип.*\b|дроп шипинг\b": "",
    r"\bтестер.*\b": "",
    r"\bдоставк.*\b|доставка\b|delivery\b": "",
    r"\bоруж.*\b|weapon\b|gun\b|blade\b|bullet\b|knife\b": "",
    r"\bлазер.*\b|laser\b": "",
}

# --- 3. СЛОВАРЬ ДЛЯ АВТОМАТИЧЕСКОГО ПРИСВОЕНИЯ КОДОВ (ваш код) ---
AUTO_CODES = {
    "духи": "3303001000", "помада": "3304100000", "для губ": "3304100000", "карандаш для глаз": "3304200000",
    "подводка": "3304200000", "зубная паста": "3306100000", "зубная нить": "3306200000", "патч для глаз": "3304990000",
    "мыло": "3401110009", "полироль": "3405300000",

    "чемодан": "4202125001", "альбом": "4820500000", "бумаг для выпечк": "4806200000", "блокнот": "4820103000",

    "шнурки": "5604100000", "москитная сетк": "5608199000", "гобелен": "5805000000",

    "шарф": "6117100000", "колготки": "6115290000", "носки": "6115950000", "фартук": "6211421000",
    "купальник": "6211120000", "бюстгальтер": "6212109000",
    "полотенце": "6302600000", "матрас надувной": "6306400000",
    "резиновые сапоги": "6401921000", "чехлы на обувь": "6401921000",
    "сабо": "6402993900", "босоножки": "6402999100", "шлепанцы": "6402999100", "сандали": "6402999100",
    "тапочк": "6405209100", "стельк": "6406905000", "мокасины": "6402999100", "слипоны": "6402999100",
    "кепка": "6505003000", "бейсболка": "6505003000", "зонт": "6601999000", "трость": "6602000000",
    "искусственн цвет": "6702100000", "парик": "6704110000",

    "внешний аккумулятор": "8507600000","воздуходувка": "8467292000","лампа автомобильная": "8539520009",
    "лечебное средство": "0","тушь для ресниц": "3304200000","зубная щетка": "9603210000","пароочиститель": "8424300100",
    "картина": "4911910000","ковер": "5705008000","чистящее средство": "3402500000","шарик воздушный": "9505900000",
    "чулки": "6115301900", "цепь для пилы": "8202400000", "цепь велосипедная": "7315111009", "катушка": "9507900000",
    "ложка чайная": "8215991000","пассатижи": "8203200009","подвеска": "7117900000","боди": "6114300000",
    "игральные карты": "9504400000","туника": "6117100000","ролики": "9506703000","термос": "9617000001",
    "рычаг для мототехники": "8714999009","вилка": "8215991000",

    "пинцет": "8203200001", "карбюратор": "8409910008", "водяной насос": "8413810000", "вентилятор": "8414592000",
    "клавиатур": "8471606000", "обратный клапан": "8481309908", "шаровой кран": "8481808199", "фен": "8516310009",
    "утюг": "8516400000", "беспроводная колонка": "8519899009", "коляск": "8715001000", "тележк": "8716800000",

    "бинокль": "9005100000", "лупа": "9013800000", "компас": "9014100000", "линейка": "9017801000",
    "часы наручные": "9102210000", "свисток": "9208900000", "спальный мешок": "9404300000", "кукла": "9503002100",
    "конструктор": "9503003500", "мягкая игрушка": "9503004100", "настольная игра": "9504908009",
    "расческа": "9615110000", "термокружка": "9617000001",

    "резинка для волос": "9615900000", "заколка для волос": "9615900000", "ершик": "9603909900",
    "гольфы": "6115961000","зеркало заднего вида": "7009100009","настенные часы": "9105210000","шапка": "6506999090",
    "поильник": "3923301090","румяна": "3304990000","сандалии": "6402993900","спонж косметический": "9616200000",
    "тени для век": "3304200000","чехол на мебель": "6304990000","электробритва": "8510100000","булавка": "7319400000",
    "выпрямитель для волос": "8516320000","плойка": "8516320000","временная татуировка": "4908900000",
    "гетры": "6406909000", "гамаши": "6406909000","мотокофр": "4202921100","пластинка виниловая": "8523809900",
    "шампунь для волос": "3305100000","лак для волос": "3305300000","чокер": "7117900000","флаг": "6307909800",
    "клетка для животных": "7326200001","шарм": "7117900000","туалетная вода": "3303009000","проектор": "8528699000",
    "пластилин": "3407000000","набор для вышивания": "6308000000","аудиокассета": "8523299000","мотоочки": "9004109100",
    "маска карнавальная": "9505900000","карнавальная": "9505900000","маска косметическая": "3304990000",
    "крючок рыболовный": "9507209000", "рыболов": "9507900000", "зарядное устройство": "8504405500",
    "звонок дверной": "8531809500", "зеркало интерьерное": "7009920000","искусственные цветы": "6702100000",
    "стабилизированные цветы": "0", "витамины": "0","карандаш": "9609109000","точилка для карандашей": "8214100000",
    "моторчик": "8501109900", "брюки": "6104630000", "свитер": "6110909000", "блузка": "6106909000",
    "юбка": "6104590000", "платье": "6104490000", "пижама": "6108390000", "перчатки": "6116990000",
    "тапочки": "6402995000", "кроссовки": "6402190000", "ботинки": "6402919000", "пирсинг": "7117900000",


    "таймер": "9106900000",  "торцевая головка": "8204200000", "уличный светильник": "9405490039",
    "фонарь": "8513100000", "футболка": "6109909000", "шумовка": "8215999000","секатор": "8201500000",
    "спортивная бутылка": "3923301090","крем": "3304990000", "веревка": "5607509000","серьги": "7117900000",
    "колье": "7117900000", "браслет": "7117900000", "кольцо": "7117900000", "комплект украшений": "7117900000",
    "свисток ": "9208900000","сборная модель": "9503003900", "рюкзак": "4202929100","розетка": "8536699008",
    "балаклава": "6505009000","галоши": "6401990000", "пульт": "8543708000", "подушка": "9404908000",
    "плед": "6301909000", "метчик": "8207401000", "одеяло": "6301909000", "аксессуар для волос": "9615900000",
    "светильник": "9405290039", "наушники": "8518309500", "корсет": "6212900000","комплект нижнего белья": "6212101000",
    "наклейки": "3919900000", "маркер": "9608200000", "защитный головной убор": "6506101000",
    "митенки": "6116930000", "весы": "8423101000", "костюм спортивный": "6112190000","бит": "8207903000",
    "галстук": "6215900000","глобус": "4905900000", "дождевик": "3926200000","дымоходная труба": "6905900000",
    "ключница настенная": "8303004000", "сейф": "8303004000", "напильник": "8203100000","ножницы": "8213000000",
    "аэрогриль": "8516607000","бинт": "3005905000", "пластырь": "3005905000", "вкладыш от пота": "4818500000",
    "гайковерт": "8467292000","дрель": "8467292000", "головоломка": "9503006900", "мозайка": "9504908009",
    "картина по номерам": "9504908009", "музыкальный диск": "8523495900","очки ": "9004109100","пинетки": "6405209900",
    "пружина": "7320208108","пряжа": "5511300000", "сверло": "8207509000","спицы для вязания": "7319909000",
    "тент": "6306120000","аквашуз": "6404199000","динамик": "8518299600", "палатка": "6306220000",
}

# --- 4. СЛОВАРЬ ДЛЯ НОРМАЛИЗАЦИИ НАИМЕНОВАНИЙ (ваш код) ---
NORMALIZE_DICT = {
    "свитер": "свитер",
    "худи": "свитер",
    "свитшот": "свитер",
    "толстовка": "свитер",
    "джемпер": "свитер",
    "пуловер": "свитер",
    "водолазка": "свитер",
    "кардиган": "свитер",

    "галоши": "резиновые сапоги",
    "чехлы на обувь": "резиновые сапоги",

    "футболка": "футболка",
    "майка": "футболка",
    "топ": "футболка",
    "поло": "футболка",
    "лонгслив": "футболка",
    "термолонгслив": "футболка",
    "термофутболка": "футболка",

    "сарафан": "платье",

    "ночная сорочка": "пижама",
    "пеньюар": "пижама",

    "брюки": "брюки",
    "штаны": "брюки",
    "легинсы": "брюки",
    "джоггеры": "брюки",

    "кроссовки": "кроссовки",
    "кеды": "кроссовки",
    "ботинки": "ботинки",
    "сапоги": "ботинки",
    "дутики": "ботинки",
    "туфли": "туфли",

    "кубики": "конструктор",

    "пуховик": "куртка",
    "парка": "куртка",
    "плащ": "куртка",
    "бомбер": "куртка",

    "шлепанцы": "босоножки",
    "сандали": "босоножки",
    "сандалии": "босоножки",
    "мокасины": "босоножки",
    "слипоны": "босоножки",
    "балетки": "босоножки",

    "докер": "кепка",
    "кепка": "кепка",
    "бейсболка": "кепка",
    "панама": "кепка",

    "бриджи": "шорты",

    "комплект трусов": "трусы",

    "ремень": "ремень",
    "пояс": "ремень",

    "заколка": "аксессуар для волос",
    "ободок": "аксессуар для волос",
    "резинка для волос": "аксессуар для волос",
    "бигуди": "аксессуар для волос",

    "набор маркеров": "маркер",
    "набор фломастеров": "маркер",

    "шлем": "защитный головной убор",
    "каска": "защитный головной убор",
    "маска защитная": "защитный головной убор",


}

# --- 5. СЛОВАРЬ ДЛЯ ВЫПАДАЮЩИХ СПИСКОВ (вы заполните его сами) ---
# Структура: Ключевое слово в наименовании -> Список вариантов (Код - Описание)
DICT_DROPDOWN = {
    "брюки": [
        ("6103430001", "Брюк. муж. синт."),
        ("6103420001", "Брюк. муж. хлоп."),
        ("6104630000", "Брюк. жен. синт."),
        ("6104620000", "Брюк. жен. хлоп."),
        ("6209200000", "Одежд. дет. хлоп."),
        ("6209300000", "Одежд. дет. синт."),
        ("6111209000", "Малыши хлоп."),
        ("6111309000", "Малыши синт.")
    ],
    "свитер": [
        ("6110309100", "муж. синт."),
        ("6110209100", "муж. хлоп."),
        ("6110113000", "муж. шерс."),
        ("6110309900", "жен. синт."),
        ("6110119000", "жен. шерс."),
        ("6110209900", "жен. хлоп."),
        ("6209200000", "Одежд. дет. хлоп."),
        ("6209300000", "Одежд. дет. синт."),
        ("6111209000", "Малыши хлоп."),
        ("6111309000", "Малыши синт.")
    ],
    # Добавьте сюда другие группы товаров
}

STOPWORDS = {"и", "в", "на", "с", "по", "к", "у", "от",
             "из"}


def normalize_name(text: str) -> str:
    """Приводит наименование к базовому виду.
    Сохраняет количество только если указана единица измерения.
    """

    if pd.isna(text):
        return ""

    # удаляем запрещённые слова
    for pattern, replacement in FORBIDDEN_WORDS.items():
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)

    text_lower = text.lower()

    # нормализация названий
    for word, base_name in NORMALIZE_DICT.items():
        text_lower = re.sub(rf'\b{re.escape(word)}\b', base_name, text_lower)

    all_words = re.findall(r"[а-яa-z0-9]+", text_lower)

    quantity_units = {
        "шт", "штук",
        "пара", "пар",
        "мл"
    }

    main_words = []
    quantity_words = []

    i = 0
    while i < len(all_words):
        word = all_words[i]

        # если число
        if word.isdigit():
            # сохраняем только если дальше единица измерения
            if i + 1 < len(all_words) and all_words[i + 1] in quantity_units:
                quantity_words.append(word)
                quantity_words.append(all_words[i + 1])
                i += 2
                continue
            else:
                # одиночную цифру удаляем
                i += 1
                continue

        if word not in quantity_units:
            main_words.append(word)

        i += 1

    clean_main_words = [w for w in main_words if w not in STOPWORDS and len(w) > 2]

    result_main = clean_main_words[:3]

    final_name = " ".join(result_main)

    if quantity_words:
        final_name += " " + " ".join(quantity_words)

    # удаление дублей подряд
    parts = final_name.split()
    deduped = []
    for word in parts:
        if not deduped or deduped[-1] != word:
            deduped.append(word)

    return " ".join(deduped).capitalize()


def add_dropdown_to_cell(wb, ws, cell, keyword, options):
    hidden_sheet_name = "_dropdowns"

    if hidden_sheet_name not in wb.sheetnames:
        hidden_ws = wb.create_sheet(hidden_sheet_name)
        hidden_ws.sheet_state = "hidden"
    else:
        hidden_ws = wb[hidden_sheet_name]

    col = None
    for c in range(1, hidden_ws.max_column + 1):
        if hidden_ws.cell(row=1, column=c).value == keyword:
            col = c
            break

    if col is None:
        col = hidden_ws.max_column + 1
        hidden_ws.cell(row=1, column=col).value = keyword

        for row_num, (code, desc) in enumerate(options, start=2):
            hidden_ws.cell(row=row_num, column=col).value = str(code)

    col_letter = hidden_ws.cell(row=2, column=col).column_letter

    # ТОЛЬКО ЛАТИНИЦА
    range_name = f"LIST_{col}"

    range_ref = f"'{hidden_sheet_name}'!${col_letter}$2:${col_letter}${len(options)+1}"

    if range_name in wb.defined_names:
        del wb.defined_names[range_name]

    wb.defined_names[range_name] = DefinedName(
        range_name,
        attr_text=range_ref
    )

    dv = DataValidation(
        type="list",
        formula1=f"={range_name}",
        allow_blank=True
    )

    ws.add_data_validation(dv)
    dv.add(cell)

    cell.value = None

    cell.fill = PatternFill(
        fill_type="solid",
        start_color="FFF2CC",
        end_color="FFF2CC"
    )

def find_tnved_code(name: str) -> str:
    """Ищет код ТНВЭД по наименованию."""
    name_lower = str(name).lower()
    for word, auto_code in AUTO_CODES.items():
        if word in name_lower:
            return auto_code
    return ""

def process_file_with_normalization(input_path: str) -> str:
    """Режим 1: Упрощает наименование + проставляет коды."""
    try:
        df = pd.read_excel(input_path, sheet_name=None)
        sheet_name = next(
            (name for name, data in df.items() if any('опис' in str(col).lower() for col in data.columns)), None)
        if not sheet_name: raise Exception("Лист с 'Описанием' не найден.")

        df_sheet = df[sheet_name]
        desc_col = next((c for c in df_sheet.columns if 'опис' in str(c).lower()), None)
        code_col = next((c for c in df_sheet.columns if 'тнвэд' in str(c).lower()), None)

        if not desc_col or not code_col:
            raise Exception("Не найдены необходимые колонки: 'Описание' и 'Тнвэд'")

        # Упрощаем наименование
        df_sheet['Описание_Новое'] = df_sheet[desc_col].apply(normalize_name)

        # Ищем коды по упрощенному наименованию
        codes = [find_tnved_code(name) for name in df_sheet['Описание_Новое'].astype(str).str.lower()]
        df_sheet['Тнвэд_Новое'] = codes

        template_path = os.path.join(os.path.dirname(__file__), "Шаблон.xlsm")
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        ws = wb.active
        source_wb = openpyxl.load_workbook(input_path)
        source_ws = source_wb.active

        for row in source_ws.iter_rows():
            for cell in row:
                ws[cell.coordinate].value = cell.value

        header = [cell.value for cell in ws[1]]
        try:
            desc_cell_idx = header.index(desc_col) + 1
            code_cell_idx = header.index(code_col) + 1
        except ValueError:
            raise Exception("Не удалось найти точное положение колонок в файле.")

        # Записываем новые значения в файл
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                row_idx = i - 2
                if row_idx >= len(df_sheet): break

                new_name = df_sheet.at[row_idx, 'Описание_Новое']
                new_code = df_sheet.at[row_idx, 'Тнвэд_Новое']

                ws.cell(row=i, column=desc_cell_idx).value = new_name

                dropdown_added = False

                for keyword, options in DICT_DROPDOWN.items():
                    if keyword in str(new_name).lower():
                        code_cell = ws.cell(row=i, column=code_cell_idx)
                        add_dropdown_to_cell(wb, ws, code_cell, keyword, options)
                        dropdown_added = True
                        break

                if not dropdown_added:
                    if str(new_code).isdigit():
                        ws.cell(row=i, column=code_cell_idx).value = int(new_code)
                    else:
                        ws.cell(row=i, column=code_cell_idx).value = ""

            except (IndexError, KeyError):
                continue


        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_norm_result.xlsm"
        wb.save(output_path)  # Сохраняем ОДИН раз в самом конце

        return output_path

    except Exception as e:
        raise Exception(f"Ошибка: {e}")


def process_file_without_normalization(input_path: str) -> str:
    """Режим 2: Только проставляет коды (наименование не трогает)."""
    try:
        df = pd.read_excel(input_path, sheet_name=None)
        sheet_name = next(
            (name for name, data in df.items() if any('опис' in str(col).lower() for col in data.columns)), None)
        if not sheet_name: raise Exception("Лист с 'Описанием' не найден.")

        df_sheet = df[sheet_name]
        desc_col = next((c for c in df_sheet.columns if 'опис' in str(c).lower()), None)
        code_col = next((c for c in df_sheet.columns if 'тнвэд' in str(c).lower()), None)

        if not desc_col or not code_col:
            raise Exception("Не найдены необходимые колонки: 'Описание' и 'Тнвэд'")

        # Ищем коды по ИСХОДНОМУ наименованию (без изменений)
        codes = [find_tnved_code(name) for name in df_sheet[desc_col].astype(str).str.lower()]
        df_sheet['Тнвэд_Новое'] = codes

        template_path = os.path.join(os.path.dirname(__file__), "Шаблон.xlsm")
        wb = openpyxl.load_workbook(template_path, keep_vba=True)
        ws = wb.active
        source_wb = openpyxl.load_workbook(input_path)
        source_ws = source_wb.active

        for row in source_ws.iter_rows():
            for cell in row:
                ws[cell.coordinate].value = cell.value

        header = [cell.value for cell in ws[1]]
        try:
            code_cell_idx = header.index(code_col) + 1
            desc_cell_idx_for_check = header.index(desc_col) + 1
        except ValueError:
            raise Exception("Не удалось найти точное положение колонок в файле.")

        # Записываем новые коды
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                row_idx = i - 2
                if row_idx >= len(df_sheet): break

                original_name = df_sheet.at[row_idx, desc_col]
                new_code = df_sheet.at[row_idx, 'Тнвэд_Новое']

                dropdown_added = False

                for keyword, options in DICT_DROPDOWN.items():
                    if keyword in str(original_name).lower():
                        code_cell = ws.cell(row=i, column=code_cell_idx)

                        add_dropdown_to_cell(wb, ws, code_cell, keyword, options)

                        dropdown_added = True
                        break

                if not dropdown_added:
                    if str(new_code).isdigit():
                        ws.cell(row=i, column=code_cell_idx).value = int(new_code)

            except (IndexError, KeyError):
                continue


        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_codes_only_result.xlsm"
        wb.save(output_path)  # Сохраняем ОДИН раз в самом конце

        return output_path

    except Exception as e:
        raise Exception(f"Ошибка: {e}")


# --- Логика GUI ---
selected_file = {"path": None}
file_label = None


def start_processing():
    if not selected_file["path"]:
        messagebox.showwarning("Ошибка", "Выберите Excel-файл")
        return
    threading.Thread(target=run_processing_with_norm, daemon=True).start()


def start_processing_codes_only():
    if not selected_file["path"]:
        messagebox.showwarning("Ошибка", "Выберите Excel-файл")
        return
    threading.Thread(target=run_processing_codes_only_thread, daemon=True).start()


def run_processing_with_norm():
    try:
        output_path = process_file_with_normalization(selected_file["path"])
        messagebox.showinfo("Готово", f"Файл обработан (с упрощением): {os.path.basename(output_path)}")
        os.startfile(os.path.dirname(output_path))
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось обработать файл:\n{e}")


def run_processing_codes_only_thread():
    try:
        output_path = process_file_without_normalization(selected_file["path"])
        messagebox.showinfo("Готово", f"Коды проставлены: {os.path.basename(output_path)}")
        os.startfile(os.path.dirname(output_path))
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось обработать файл:\n{e}")


def choose_file():
    path = filedialog.askopenfilename(
        title="Выберите Excel файл",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )
    if path:
        selected_file["path"] = path
        file_label.config(text=os.path.basename(path))


root = tk.Tk()
root.title("Обработка ТН ВЭД")
root.geometry("450x250")
root.resizable(False, False)
ttk.Label(root, text="Выберите Excel-файл для обработки:").pack(pady=10)

file_label = ttk.Label(root, text="Файл не выбран")
file_label.pack(pady=5)

ttk.Button(root, text="Выбрать файл", command=choose_file).pack(pady=5)

# Две кнопки для разных режимов работы
frame = ttk.Frame(root)
frame.pack(pady=10)

ttk.Button(frame, text="Обработать (с упрощением)",
            command=start_processing).pack(side=tk.LEFT, padx=5)

ttk.Button(frame, text="Проставить коды (без упрощения)",
            command=start_processing_codes_only).pack(side=tk.LEFT, padx=5)

root.mainloop()