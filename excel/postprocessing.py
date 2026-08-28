from openpyxl.styles import PatternFill, Font

from dictionaries.all_dictionaries import (
    ALLOWED_PREFIXES,
    RESTRICTED_PREFIXES,
)


# ============================================================
# ЦВЕТА СТАТУСА СТРОКИ
# ============================================================
RED_ROW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FCE4E4",
)
GREEN_ROW_FILL = PatternFill(
    fill_type="solid",
    fgColor="E8F5E9",
)
RED_FONT = Font(
    color="C00000",
)
GREEN_FONT = Font(
    color="008000",
)
# ============================================================
# НОРМАЛИЗАЦИЯ КОДА
# ============================================================
def _normalize_code(value):
    """
    Приводит значение к нормальной строке кода.

    """

    if value is None:
        return ""

    if isinstance(value, (int, float)):
        if value == 0:
            return ""

        if isinstance(value, float) and value.is_integer():
            value = int(value)

    code = str(value).strip()

    if code in ("0", "0.0", "0,0"):
        return ""

    code = code.replace(" ", "")

    return code
# ============================================================
# ПРОВЕРКИ
# ============================================================
def _is_zero_code(value):
    """
    Проверяет, является ли значение нулевым/пустым кодом.
    """

    if value is None:
        return True

    if isinstance(value, (int, float)):
        return value == 0

    value = str(value).strip().replace(" ", "")

    return value in (
        "",
        "0",
        "0.0",
        "0,0",
        "None",
        "nan",
        "NaN",
    )


def _is_allowed(code):
    """
    Разрешённое исключение.
    """

    if not code:
        return False

    return any(
        code.startswith(prefix)
        for prefix in ALLOWED_PREFIXES
    )


def _is_restricted(code):
    """
    Запрещённый код.
    """

    if not code:
        return False

    return any(
        code.startswith(prefix)
        for prefix in RESTRICTED_PREFIXES
    )
# ============================================================
# ОСНОВНАЯ ПОСТОБРАБОТКА
# ============================================================
def apply_visual_postprocessing(
    ws,
    code_col_idx,
    decision_col_idx=None,
    max_row=None,
):
    """
    Финальная визуальная обработка Excel.
    Логика:
        пустой / 0
            ↓
        КРАСНАЯ СТРОКА
        ALLOWED_PREFIXES
            ↓
        ЗЕЛЁНАЯ СТРОКА
        RESTRICTED_PREFIXES
            ↓
        КРАСНАЯ СТРОКА
        обычный код
            ↓
        ничего не меняем
    """

    if max_row is None:
        max_row = ws.max_row

    red_count = 0
    green_count = 0
    zero_count = 0
    restricted_count = 0
    normal_count = 0

    print(
        "\n"
        "============================================================\n"
        "VISUAL POSTPROCESSING\n"
        "============================================================"
    )
    print(f"Worksheet: {ws.title}")

    print(f"Rows: 2 -> {max_row}")

    print(f"Code column: {code_col_idx}")

    for row in range(2, max_row + 1):

        code_cell = ws.cell(
            row=row,
            column=code_col_idx,
        )
        raw_value = code_cell.value
        code = _normalize_code(raw_value)
        # ----------------------------------------------------
        # 1. ПУСТОЙ / 0
        # ----------------------------------------------------
        if _is_zero_code(raw_value):

            _paint_row(
                ws=ws,
                row=row,
                fill=RED_ROW_FILL,
                font=RED_FONT,
                code_col_idx=code_col_idx,
            )
            if decision_col_idx:
                ws.cell(
                    row=row,
                    column=decision_col_idx,
                ).value = "Нельзя"
            red_count += 1
            zero_count += 1
            print(
                f"[RED][ZERO] row={row} "
                f"raw={raw_value!r}"
            )
            continue
        # ----------------------------------------------------
        # 2. РАЗРЕШЁННОЕ ИСКЛЮЧЕНИЕ
        # ----------------------------------------------------
        if _is_allowed(code):

            _paint_row(
                ws=ws,
                row=row,
                fill=GREEN_ROW_FILL,
                font=GREEN_FONT,
                code_col_idx=code_col_idx,
            )
            if decision_col_idx:
                ws.cell(
                    row=row,
                    column=decision_col_idx,
                ).value = "Можно"
            green_count += 1
            print(
                f"[GREEN][ALLOWED] row={row} "
                f"code={code}"
            )
            continue
        # ----------------------------------------------------
        # 3. ЗАПРЕЩЁННЫЙ КОД
        # ----------------------------------------------------
        if _is_restricted(code):

            _paint_row(
                ws=ws,
                row=row,
                fill=RED_ROW_FILL,
                font=RED_FONT,
                code_col_idx=code_col_idx,
            )
            if decision_col_idx:
                ws.cell(
                    row=row,
                    column=decision_col_idx,
                ).value = "Нельзя"
            red_count += 1
            restricted_count += 1
            print(
                f"[RED][RESTRICTED] row={row} "
                f"code={code}"
            )
            continue
        # ----------------------------------------------------
        # 4. ОБЫЧНЫЙ КОД
        # ----------------------------------------------------
        normal_count += 1

        if decision_col_idx:
            ws.cell(
                row=row,
                column=decision_col_idx,
            ).value = "Можно"
    print(
        "\n"
        "------------------------------------------------------------"
    )
    print(f"ZERO:       {zero_count}")
    print(f"RESTRICTED: {restricted_count}")
    print(f"RED TOTAL:  {red_count}")
    print(f"GREEN:      {green_count}")
    print(f"NORMAL:     {normal_count}")
    print(
        "============================================================\n"
    )
    return {
        "red": red_count,
        "green": green_count,
        "zero": zero_count,
        "restricted": restricted_count,
        "normal": normal_count,
    }
# ============================================================
# ПОКРАСКА СТРОКИ
# ============================================================
def _paint_row(
    ws,
    row,
    fill,
    font,
    code_col_idx,
):
    """
    Красит всю строку, кроме ячейки кода.
    Ячейку кода пропускаем специально,
    потому что её цвет отвечает за MATERIAL_COLORS.
    """
    for col in range(
        1,
        ws.max_column + 1,
    ):
        # Не трогаем цвет кода
        if col == code_col_idx:
            continue

        cell = ws.cell(
            row=row,
            column=col,
        )
        cell.fill = fill
        cell.font = font