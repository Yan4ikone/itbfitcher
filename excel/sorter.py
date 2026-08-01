from copy import copy

from openpyxl.formula.translate import Translator


def sort_by_description(ws, desc_col_idx, last_row):

    rows_data = []
    for row in range(2, last_row + 1):
        desc_value = ws.cell(row=row, column=desc_col_idx).value
        if not desc_value:
            continue
        row_cells = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            row_cells.append({
                "coordinate": cell.coordinate,
                "value": cell.value,
                "font": copy(cell.font) if cell.has_style else None,
                "fill": copy(cell.fill) if cell.has_style else None,
                "alignment": copy(cell.alignment) if cell.has_style else None,
                "border": copy(cell.border) if cell.has_style else None,
                "number_format": cell.number_format if cell.has_style else None,
            })
        rows_data.append(row_cells)

    if not rows_data:
        return

    desc_idx = desc_col_idx - 1
    rows_data.sort(key=lambda r: str(r[desc_idx]["value"] or "").lower())

    # Перезаписываем строки в отсортированном порядке
    for new_row_idx, row_cells in enumerate(rows_data, start=2):
        for col_idx, cell_data in enumerate(row_cells, start=1):
            cell = ws.cell(row=new_row_idx, column=col_idx)
            value = cell_data["value"]

            if isinstance(value, str) and value.startswith("="):
                try:
                    value = Translator(
                        value,
                        origin=cell_data["coordinate"]
                    ).translate_formula(
                        cell.coordinate
                    )
                except Exception:
                    pass
            cell.value = value
            if cell_data["font"]:
                cell.font = cell_data["font"]
            if cell_data["fill"]:
                cell.fill = cell_data["fill"]
            if cell_data["alignment"]:
                cell.alignment = cell_data["alignment"]
            if cell_data["border"]:
                cell.border = cell_data["border"]
            if cell_data["number_format"]:
                cell.number_format = cell_data["number_format"]
