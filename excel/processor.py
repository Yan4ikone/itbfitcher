import os
import openpyxl

from classifier.decision_engine import DecisionEngine
from excel.reader import load_input, load_source_sheet
from excel.utils import copy_sheet
from excel.writer import (write_result, add_history_warning, fill_surname_column, save_workbook)
from excel.sorter import sort_by_description
from excel.template import (TEMPLATE_PATH, detect_template_structure)
from excel.restrictions import (apply_restrictions, apply_description_warnings)
from learning.importer import load_learning_history
from models.card_builder import build_from_excel
from modules.classification_statistics import (ClassificationStatistics)
from modules.dropdown_manager import (apply_specific_dropdowns)
from modules.result_engine import (apply_result)
from classifier.normalizer import (normalize_name)


def _log(text, logger=None,):

    if logger:
        logger(text)
    else:
        print(text)


def process_file_with_normalization(
    input_path,
    logger=None,
    progress_callback=None,
):

    learning_history = load_learning_history(input_path)
    data = load_input(input_path)
    df = data["dataframe"]
    sheet_name = data["sheet_name"]
    desc_col = data["columns"]["description"]
    code_col = data["columns"]["code"]
    characteristics_col = data["columns"]["characteristics"]
    total_rows = data["rows"]

    df["Описание_Новое"] = (
        df[desc_col]
        .fillna("")
        .apply(normalize_name)
    )

    last_row = len(df) + 1

    if not os.path.exists(TEMPLATE_PATH):
        raise Exception(
            f"Шаблон не найден: {TEMPLATE_PATH}"
        )

    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb.active

    _log("Копирование шаблона...", logger)
    source_ws = load_source_sheet(input_path)
    copy_sheet(source_ws, ws)
    structure = detect_template_structure(ws)
    desc_cell_idx = structure["desc_col"]
    code_cell_idx = structure["code_col"]
    decision_col_idx = structure["decision_col"]
    surname_col_idx = structure["surname_col"]
    stats = ClassificationStatistics()
    engine = DecisionEngine(learning_history)
    processed_rows = 0
    _log(f"Всего строк: {total_rows}", logger)

    for row in range(2, last_row + 1):

        try:
            dataframe_row = row - 2
            original_name = str(
                df.at[
                    dataframe_row,
                    desc_col,
                ]
            ).strip()

            if (
                not original_name
                or original_name.lower()
                == "nan"
            ):
                continue
            normalized_name = str(
                df.at[
                    dataframe_row,
                    "Описание_Новое",
                ]
            )
            characteristics = ""

            if characteristics_col:
                characteristics = str(
                    df.at[
                        dataframe_row,
                        characteristics_col,
                    ]
                )

            card = build_from_excel(
                description=normalized_name,
                characteristics=characteristics,
            )

            result = engine.decide(card)

            history = learning_history.get(
                normalized_name.lower()
            )

            add_history_warning(
                ws=ws,
                row=row,
                code_column=code_cell_idx,
                history=history,
            )

            stats.add(result)

            write_result(
                ws=ws,
                row=row,
                code_column=code_cell_idx,
                description_column=desc_cell_idx,
                description=normalized_name,
                result=result,
                apply_result=apply_result,
            )

            processed_rows += 1

            if (
                    processed_rows % 10 == 0
                    or processed_rows == total_rows
            ):

                if progress_callback:
                    progress_callback(
                        total_rows,
                        processed_rows,
                    )

                _log(
                    f"Обработано "
                    f"{processed_rows}/"
                    f"{total_rows}",
                    logger,
                )

        except Exception as error:

            _log(
                f"Ошибка строки "
                f"{row}: {error}",
                logger,
            )

    fill_surname_column(
        ws,
        surname_col_idx,
    )

    _log(
        "Сортировка...",
        logger,
    )

    sort_by_description(
        ws,
        desc_cell_idx,
        last_row,
    )

    _log(
        "Выпадающие списки...",
        logger,
    )

    apply_specific_dropdowns(
        ws,
        desc_cell_idx,
        code_cell_idx,
        max_row=last_row,
    )

    _log(
        "Проверка ограничений...",
        logger,
    )

    apply_restrictions(
        ws,
        code_cell_idx,
        decision_col_idx,
        surname_col_idx,
        is_first_pass=True,
        max_row=last_row,
    )

    apply_description_warnings(
        ws,
        desc_cell_idx,
        max_row=last_row,
    )

    base, _ = os.path.splitext(
        input_path
    )

    output_path = (
        f"{base}_norm_result.xlsm"
    )

    _log(
        "Сохранение...",
        logger,
    )

    save_workbook(
        wb,
        output_path,
    )

    stats.print_summary(
        logger
    )

    stats.save_excel(
        input_path
    )

    if progress_callback:
        progress_callback(
            total_rows,
            total_rows,
        )

    _log(
        f"Готово: "
        f"{os.path.basename(output_path)}",
        logger,
    )

    return output_path

def recalculate_codes(input_path, logger=None, progress_callback=None,):

    learning_history = load_learning_history(input_path)

    data = load_input(input_path)
    df = data["dataframe"]
    desc_col = data["columns"]["description"]
    code_col = data["columns"]["code"]
    characteristics_col = data["columns"]["characteristics"]
    total_rows = data["rows"]
    last_row = len(df) + 1

    if not os.path.exists(TEMPLATE_PATH):
        raise Exception(f"Шаблон не найден: {TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb.active
    source_wb = openpyxl.load_workbook(input_path)
    source_ws = source_wb.active

    _log(
        "Копирование шаблона...",
        logger,
    )

    copy_sheet(source_ws, ws)
    header = [
        cell.value
        for cell in ws[1]
    ]

    try:

        desc_cell_idx = (header.index(desc_col) + 1)
        code_cell_idx = (header.index(code_col) + 1)

    except ValueError:

        raise Exception("Не удалось определить колонки.")

    status_column = None
    for column in df.columns:
        text = str(column).lower()
        if any(
            word in text
            for word in (
                "статус",
                "огранич",
                "можно",
                "разреш",
            )

        ):

            status_column = column
            break

    if status_column:

        status_cell_idx = (header.index(status_column) + 1)

    else:

        status_cell_idx = None

    engine = DecisionEngine(learning_history)
    processed_rows = 0

    _log(f"Всего строк: {total_rows}", logger)

    for row in range(2, last_row + 1):

        try:

            description = ws.cell(
                row=row,
                column=desc_cell_idx,
            ).value

            if not description:
                continue

            characteristics = ""

            if characteristics_col:
                characteristics = str(
                    df.at[
                        row - 2,
                        characteristics_col,
                    ]
                )

            card = build_from_excel(
                description=str(description),
                characteristics=characteristics,
            )

            result = engine.decide(card)
            processed_rows += 1

            if (
                processed_rows % 10 == 0
                or processed_rows == total_rows
            ):

                if progress_callback:

                    progress_callback(
                        total_rows,
                        processed_rows,
                    )

                _log(
                    f"Пересчитано "
                    f"{processed_rows}/"
                    f"{total_rows}",
                    logger,
                )

        except Exception as error:

            _log(
                f"Ошибка строки "
                f"{row}: {error}",
                logger,
            )

    apply_restrictions(
        ws,
        code_cell_idx,
        status_cell_idx,
        None,
        is_first_pass=False,
        max_row=last_row,
    )

    apply_description_warnings(
        ws,
        desc_cell_idx,
        max_row=last_row,
    )

    base, _ = os.path.splitext(input_path)
    output_path = (f"{base}_recalc_result.xlsm")
    save_workbook(wb, output_path)

    if progress_callback:

        progress_callback(
            total_rows,
            total_rows,
        )

    _log(
        f"Готово: "
        f"{os.path.basename(output_path)}",
        logger,
    )

    return output_path