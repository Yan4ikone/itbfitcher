from openpyxl.styles import Font

from dictionaries.all_dictionaries import RESTRICTED_PREFIXES, FORBIDDEN_TRIGGERS
from excel.styles import RED_FONT, WARNING_FILL


def apply_restrictions(ws, code_col_idx, decision_col_idx=None, surname_col_idx=None, is_first_pass=True, max_row=None):
    limit = max_row if max_row else ws.max_row
    for row in range(2, limit + 1):
        code_cell = ws.cell(row=row, column=code_col_idx)
        decision_cell = ws.cell(row=row, column=decision_col_idx) if decision_col_idx else None

        code_val = str(code_cell.value).strip() if code_cell.value else ""
        last_col = ws.max_column

        while last_col > 1 and ws.cell(1, last_col).value is None:
            last_col -= 1

        # 1. Код "0" или пустой -> Красный + Нельзя
        if not code_val or code_val == "0":
            code_cell.font = RED_FONT
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).font = RED_FONT
            if decision_cell: decision_cell.value = "Нельзя"
            continue

        # 2. ПЕРВИЧНАЯ ОБРАБОТКА: Просто ставим "Можно", не красим
        if is_first_pass:
            code_cell.font = Font(color="000000")
            if decision_cell: decision_cell.value = "Можно"
            continue

        # 3. ПОВТОРНАЯ ПРОВЕРКА: Проверяем на запреты -> красим ВСЮ СТРОКУ
        is_restricted = any(code_val.startswith(p) for p in RESTRICTED_PREFIXES)
        if is_restricted:
            code_cell.font = RED_FONT
            for col in range(1, last_col + 1):
                ws.cell(row=row, column=col).font = RED_FONT
            if decision_cell: decision_cell.value = "Нельзя"
            continue

        # Остальные коды -> Черный + Можно
        code_cell.font = Font(color="000000")
        if decision_cell: decision_cell.value = "Можно"

def apply_description_warnings(ws, desc_col_idx, max_row=None):
    limit = max_row if max_row else ws.max_row
    for row in range(2, limit + 1):
        cell = ws.cell(row=row, column=desc_col_idx)
        if not cell.value:
            continue
        text = str(cell.value).lower()
        for trigger in FORBIDDEN_TRIGGERS:
            if trigger in text:
                cell.fill = WARNING_FILL  # Жёлтый фон
                break