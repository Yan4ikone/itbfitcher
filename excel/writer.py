from openpyxl.comments import Comment

from excel.styles import WARNING_FILL


def write_result(
        ws,
        row,
        code_column,
        description_column,
        description,
        result,
        apply_result,
):

    apply_result(
        ws,
        row,
        code_column,
        description,
        result,
    )

    ws.cell(
        row=row,
        column=description_column,
    ).value = description


def add_history_warning(
        ws,
        row,
        code_column,
        history,
):

    if (
        not history
        or len(history["codes"]) <= 1
    ):
        return

    cell = ws.cell(
        row=row,
        column=code_column,
    )

    cell.fill = WARNING_FILL

    variants = [
        code

        for code, _ in history[
            "codes"
        ].most_common(10)

    ]

    text = ("В истории встречались разные коды:\n" + "\n".join(variants))

    if cell.comment:

        cell.comment.text += ( "\n\n" + text)

    else:

        cell.comment = Comment(text, "AutoClassifier")


def fill_surname_column(ws, surname_column):

    if not surname_column:
        return

    for row in range(2, ws.max_row + 1):

        ws.cell(row=row, column=surname_column).value = (f'=IF(C{row}<>"","Куратов","")')


def save_workbook(workbook, output_path):

    workbook.save(output_path)